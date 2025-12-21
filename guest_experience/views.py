from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.db.models import Count, Q, Avg, Min, Max
from django.db.models.functions import TruncDate
import json
import pandas as pd
import io
from datetime import datetime, timedelta
from .models import ArrivalRecord
from django.db import models

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def dashboard(request):
    """
    Guest Experience - Arrivals Dashboard with charts and metrics.
    """
    today = timezone.localdate()
    
    # Get filter parameters
    start_date = request.GET.get('start_date', (today - timedelta(days=30)).isoformat())
    end_date = request.GET.get('end_date', today.isoformat())
    status_filter = request.GET.get('status', '')
    country_filter = request.GET.get('country', '')
    travel_agent_filter = request.GET.get('travel_agent', '')
    search_query = request.GET.get('search', '')
    
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date_obj = today - timedelta(days=30)
        end_date_obj = today
        start_date = start_date_obj.isoformat()
        end_date = end_date_obj.isoformat()

    # Build base queryset for dropdown options (from current date range)
    base_qs = ArrivalRecord.objects.filter(
        arrival_date__gte=start_date_obj,
        arrival_date__lte=end_date_obj,
    )
    # Optionally respect current status/country filters when building dropdown
    if status_filter:
        base_qs = base_qs.filter(status__iexact=status_filter)
    if country_filter:
        base_qs = base_qs.filter(country__icontains=country_filter)

    travel_agents = (
        base_qs.exclude(travel_agent_name__isnull=True)
        .exclude(travel_agent_name="")
        .values_list("travel_agent_name", flat=True)
        .order_by("travel_agent_name")
        .distinct()
    )
    
    context = {
        "section": "guest_experience",
        "subsection": "dashboard",
        "page_title": "Guest Experience - Arrivals Dashboard",
        "today": today,
        "start_date": start_date,
        "end_date": end_date,
        "status_filter": status_filter,
        "country_filter": country_filter,
        "travel_agent_filter": travel_agent_filter,
        "search_query": search_query,
        "travel_agents": travel_agents,
    }
    return render(request, "guest_experience/dashboard.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def arrivals(request):
    """
    Guest Experience - Arrivals page.
    Currently uses placeholder data; wire to your PMS data later.
    """
    today = timezone.localdate()
    context = {
        "section": "guest_experience",
        "subsection": "arrivals",
        "page_title": "Guest Experience - Arrivals",
        "today": today,
    }
    return render(request, "guest_experience/arrivals.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def upload_arrivals(request):
    """
    Upload an Excel file (e.g. 'Arrival reporte.xlsx') to populate ArrivalRecord.

    Expected columns (case-insensitive):
      - Room
      - Guest Name
      - ETA
      - Nights
      - Status
      - Arrival Date
    """
    if request.method != "POST":
        return redirect("guest_experience:arrivals")

    file = request.FILES.get("file")
    if not file:
        messages.error(request, "Please choose an Excel file to upload.")
        return redirect("guest_experience:arrivals")

    try:
        # Read without assuming the first row is the header, so we can detect the real header row
        raw_df = pd.read_excel(file, header=None)
    except Exception as exc:
        messages.error(request, f"Could not read Excel file: {exc}")
        return redirect("guest_experience:arrivals")

    # Try to detect the header row by looking for a cell equal to 'Arrival Date' (case-insensitive)
    header_row_idx = None
    target_header = "arrival date"
    for idx, row in raw_df.iterrows():
        for val in row:
            if isinstance(val, str) and val.strip().lower() == target_header:
                header_row_idx = idx
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        # Fallback: treat the first row as header
        header_row_idx = 0

    header = raw_df.iloc[header_row_idx].tolist()
    df = raw_df.iloc[header_row_idx + 1 :].copy()

    # Normalize column names for flexible matching
    df.columns = [str(c).strip().lower() for c in header]

    def col(*names):
        for n in names:
            n = n.lower()
            if n in df.columns:
                return n
        return None

    # Map your Excel headers
    property_col = col("property")
    confirmation_col = col("confirmation number", "confirmation")
    first_name_col = col("first name", "firstname", "first")
    last_name_col = col("last name", "lastname", "last")
    room_col = col("room number", "room")
    phone_col = col("phone", "phone number", "tel")
    email_col = col("email", "e-mail")
    nationality_col = col("nationality")
    country_col = col("country")
    guest_col = col("guest name", "guest", "name")
    first_name_col = col("first name", "firstname", "first")
    last_name_col = col("last name", "lastname", "last")
    eta_col = col("eta", "arrival time")
    nights_col = col("nights", "length of stay")
    status_col = col("status")
    arrival_date_col = col("arrival date", "check in", "arrival")
    departure_date_col = col("departure date", "check out", "checkout", "departure")

    # We must at least have arrival date; room number is optional
    if not arrival_date_col:
        messages.error(
            request,
            "Excel file must contain an 'Arrival Date' column.",
        )
        return redirect("guest_experience:arrivals")

    # If we have confirmation numbers, clean existing rows for those confirmations
    # so re-importing the same Excel file does not create duplicates.
    if confirmation_col:
        confirmation_values = set()
        for _, row in df.iterrows():
            cn = str(row.get(confirmation_col) or "").strip()
            if cn:
                confirmation_values.add(cn)
        if confirmation_values:
            ArrivalRecord.objects.filter(confirmation_number__in=confirmation_values).delete()

    created_or_updated = 0
    today = timezone.localdate()
    for _, row in df.iterrows():
        arrival_date_val = row.get(arrival_date_col)
        if pd.isna(arrival_date_val):
            continue
        try:
            arrival_date = pd.to_datetime(arrival_date_val).date()
        except Exception:
            # Skip rows with bad dates
            continue

        property_name = str(row.get(property_col) or "").strip() if property_col else ""
        confirmation_number = str(row.get(confirmation_col) or "").strip() if confirmation_col else ""
        first_name = str(row.get(first_name_col) or "").strip() if first_name_col else ""
        last_name = str(row.get(last_name_col) or "").strip() if last_name_col else ""
        room = str(row.get(room_col) or "").strip() if room_col else ""
        phone = str(row.get(phone_col) or "").strip() if phone_col else ""
        email = str(row.get(email_col) or "").strip() if email_col else ""
        nationality = str(row.get(nationality_col) or "").strip() if nationality_col else ""
        country = str(row.get(country_col) or "").strip() if country_col else ""

        # Build guest name from either a combined column or First/Last Name
        if guest_col:
            guest_name = str(row.get(guest_col) or "").strip()
        else:
            first = str(row.get(first_name_col) or "").strip() if first_name_col else ""
            last = str(row.get(last_name_col) or "").strip() if last_name_col else ""
            guest_name = (first + " " + last).strip()

        eta = "" if not eta_col else str(row.get(eta_col) or "").strip()

        # Departure date (store raw date if present)
        departure_date = None
        if departure_date_col:
            dep_val_for_store = row.get(departure_date_col)
            if pd.notna(dep_val_for_store):
                try:
                    departure_date = pd.to_datetime(dep_val_for_store).date()
                except Exception:
                    departure_date = None

        # Nights: from explicit column if present, otherwise compute from Arrival/Departure
        nights = None
        if nights_col:
            nights_raw = row.get(nights_col)
            try:
                nights = int(nights_raw) if pd.notna(nights_raw) else None
            except (TypeError, ValueError):
                nights = None
        elif departure_date is not None:
            try:
                diff = (departure_date - arrival_date).days
                nights = diff if diff >= 0 else None
            except Exception:
                nights = None

        # Status: from column; default to "Expected" if blank
        raw_status = "" if not status_col else str(row.get(status_col) or "").strip()
        status = raw_status or "Expected"

        # If no confirmation number, skip to avoid untrackable duplicates
        if not confirmation_number:
            continue

        ArrivalRecord.objects.create(
            # Raw columns
            property_name=property_name,
            confirmation_number=confirmation_number,
            first_name=first_name,
            last_name=last_name,
            room=room or None,
            phone=phone,
            email=email,
            nationality=nationality,
            country=country,
            arrival_date=arrival_date,
            departure_date=departure_date,
            travel_agent_name=str(row.get(col("travel agent name", "agent", "travel agent")) or "").strip(),
            # Derived/display
            guest_name=guest_name,
            eta=eta,
            nights=nights,
            status=status,
            created_by=request.user,
            updated_by=request.user,
        )
        created_or_updated += 1

    messages.success(
        request,
        f"Imported {created_or_updated} arrival rows from Excel.",
    )
    return redirect("guest_experience:arrivals")


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def in_house(request):
    """
    Guest Experience - In-House Guests page.
    """
    today = timezone.localdate()
    context = {
        "section": "guest_experience",
        "subsection": "in_house",
        "page_title": "Guest Experience - In-House Guests",
        "today": today,
    }
    return render(request, "guest_experience/in_house.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def departures(request):
    """
    Guest Experience - Departures page.
    """
    today = timezone.localdate()
    context = {
        "section": "guest_experience",
        "subsection": "departures",
        "page_title": "Guest Experience - Departures",
        "today": today,
    }
    return render(request, "guest_experience/departures.html", context)


# --- Simple JSON API placeholders ---

@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def arrivals_api(request):
    """
    API endpoint for arrivals.
    Returns arrivals from ArrivalRecord for today (or ?date=YYYY-MM-DD).
    """
    date_param = request.GET.get("date")
    try:
        if date_param:
            target_date = timezone.datetime.strptime(date_param, "%Y-%m-%d").date()
        else:
            target_date = timezone.localdate()
    except ValueError:
        return JsonResponse({"error": "Invalid date format, expected YYYY-MM-DD."}, status=400)

    # Exclude records with "In-House" status (case-insensitive, handles both "In-House" and "in house")
    qs = ArrivalRecord.objects.filter(
        arrival_date=target_date
    ).exclude(
        models.Q(status__iexact="in-house") | models.Q(status__iexact="in house")
    ).order_by("room")
    data = []
    for a in qs:
        data.append(
            {
                "room": a.room,
                "guest_name": a.guest_name,
                "eta": a.eta,
                "nights": a.nights,
                "status": a.status,
                "property_name": a.property_name,
                "confirmation_number": a.confirmation_number,
                "first_name": a.first_name,
                "last_name": a.last_name,
                "phone": a.phone,
                "email": a.email,
                "nationality": a.nationality,
                "country": a.country,
                "arrival_date": a.arrival_date.isoformat() if a.arrival_date else None,
                "departure_date": a.departure_date.isoformat() if a.departure_date else None,
                "travel_agent_name": a.travel_agent_name,
            }
        )
    return JsonResponse({"results": data})


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def edit_arrival(request, confirmation_number):
    """
    Edit an arrival record by confirmation number.
    GET: Returns JSON with record data
    POST: Updates the record
    """
    try:
        record = ArrivalRecord.objects.get(confirmation_number=confirmation_number)
    except ArrivalRecord.DoesNotExist:
        return JsonResponse({"error": "Record not found"}, status=404)

    if request.method == "GET":
        # Return record data for editing
        return JsonResponse({
            "confirmation_number": record.confirmation_number,
            "room": record.room or "",
            "first_name": record.first_name or "",
            "last_name": record.last_name or "",
            "guest_name": record.guest_name or "",
            "phone": record.phone or "",
            "email": record.email or "",
            "nationality": record.nationality or "",
            "country": record.country or "",
            "arrival_date": record.arrival_date.isoformat() if record.arrival_date else "",
            "departure_date": record.departure_date.isoformat() if record.departure_date else "",
            "travel_agent_name": record.travel_agent_name or "",
            "status": record.status or "",
            "eta": record.eta or "",
        })

    elif request.method == "POST":
        # Update the record
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        # Update fields
        if "room" in payload:
            record.room = payload["room"].strip() or None
        if "first_name" in payload:
            record.first_name = payload["first_name"].strip()
        if "last_name" in payload:
            record.last_name = payload["last_name"].strip()
        if "guest_name" in payload:
            record.guest_name = payload["guest_name"].strip()
        elif "first_name" in payload or "last_name" in payload:
            # Auto-update guest_name if first/last name changed
            first = record.first_name or ""
            last = record.last_name or ""
            record.guest_name = (first + " " + last).strip()
        if "phone" in payload:
            record.phone = payload["phone"].strip()
        if "email" in payload:
            record.email = payload["email"].strip()
        if "nationality" in payload:
            record.nationality = payload["nationality"].strip()
        if "country" in payload:
            record.country = payload["country"].strip()
        if "arrival_date" in payload:
            try:
                if payload["arrival_date"]:
                    record.arrival_date = timezone.datetime.strptime(payload["arrival_date"], "%Y-%m-%d").date()
                else:
                    record.arrival_date = None
            except (ValueError, TypeError):
                pass
        if "departure_date" in payload:
            try:
                if payload["departure_date"]:
                    record.departure_date = timezone.datetime.strptime(payload["departure_date"], "%Y-%m-%d").date()
                else:
                    record.departure_date = None
            except (ValueError, TypeError):
                pass
        if "travel_agent_name" in payload:
            record.travel_agent_name = payload["travel_agent_name"].strip()
        if "status" in payload:
            record.status = payload["status"].strip()
        if "eta" in payload:
            record.eta = payload["eta"].strip()

        # Recalculate nights if dates changed
        if record.arrival_date and record.departure_date:
            try:
                diff = (record.departure_date - record.arrival_date).days
                record.nights = diff if diff >= 0 else None
            except Exception:
                pass

        record.updated_by = request.user
        record.updated_at = timezone.now()
        record.save()

        return JsonResponse({
            "success": True,
            "message": "Record updated successfully",
            "confirmation_number": record.confirmation_number
        })


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
@require_POST
def delete_arrivals(request):
    """
    Bulk delete arrivals by confirmation numbers.
    Expects JSON body: {"confirmation_numbers": ["ABC123", "DEF456", ...]}
    """
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    ids = payload.get("confirmation_numbers") or []
    if not isinstance(ids, list):
        return JsonResponse({"error": "confirmation_numbers must be a list"}, status=400)

    # Normalize and filter out empty values
    cleaned_ids = [str(x).strip() for x in ids if str(x).strip()]
    if not cleaned_ids:
        return JsonResponse({"deleted": 0})

    deleted_count, _ = ArrivalRecord.objects.filter(confirmation_number__in=cleaned_ids).delete()
    return JsonResponse({"deleted": deleted_count})


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
@require_POST
def mark_in_house(request):
    """
    Bulk update arrivals to In-House status by confirmation numbers.
    Expects JSON body: {"confirmation_numbers": ["ABC123", "DEF456", ...]}
    """
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    ids = payload.get("confirmation_numbers") or []
    if not isinstance(ids, list):
        return JsonResponse({"error": "confirmation_numbers must be a list"}, status=400)

    cleaned_ids = [str(x).strip() for x in ids if str(x).strip()]
    if not cleaned_ids:
        return JsonResponse({"updated": 0})

    now = timezone.now()
    updated = 0
    qs = ArrivalRecord.objects.filter(confirmation_number__in=cleaned_ids)
    for record in qs:
        # Only set in-house timing if actually transitioning to In-House
        if (record.status or "").lower() not in ["in-house", "in house"]:
            record.status = "In-House"
            record.in_house_since = now
            record.in_house_by = request.user

            # First courtesy call: 20 minutes after status change
            record.first_courtesy_due_at = now + timedelta(minutes=20)

            # Second courtesy call: next day if stay > 1 night and before departure
            second_due = None
            if record.nights and record.nights > 1:
                candidate = now + timedelta(days=1)
                if record.departure_date:
                    # Only schedule if still before departure date
                    if candidate.date() <= record.departure_date:
                        second_due = candidate
                else:
                    second_due = candidate
            record.second_courtesy_due_at = second_due

        record.updated_by = request.user
        record.updated_at = now
        record.save()
        updated += 1

    return JsonResponse({"updated": updated})


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
@require_POST
def mark_departed(request):
    """
    Mark a guest as departed from In-House status.
    Expects JSON body: {
        "confirmation_number": "ABC123",
        "departure_method": "Checkout at front desk",
        "departure_notes": "Guest was satisfied",
        "message_sent_to_guest": true
    }
    """
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    confirmation_number = str(payload.get("confirmation_number") or "").strip()
    departure_method = str(payload.get("departure_method") or "").strip()
    departure_notes = str(payload.get("departure_notes") or "").strip()
    message_sent_to_guest = bool(payload.get("message_sent_to_guest", False))

    if not confirmation_number:
        return JsonResponse({"error": "confirmation_number is required"}, status=400)
    if not departure_method:
        return JsonResponse({"error": "departure_method is required"}, status=400)

    try:
        record = ArrivalRecord.objects.get(confirmation_number=confirmation_number)
    except ArrivalRecord.DoesNotExist:
        return JsonResponse({"error": "Record not found"}, status=404)

    now = timezone.now()
    record.status = "Departed"
    record.departed_at = now
    record.departed_by = request.user
    record.departure_method = departure_method
    record.departure_notes = departure_notes
    record.message_sent_to_guest = message_sent_to_guest
    record.updated_by = request.user
    record.updated_at = now
    record.save()

    return JsonResponse({"success": True})


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
@require_POST
def update_room(request):
    """
    Update the room number for an in-house guest.
    Expects JSON body: {
        "confirmation_number": "ABC123",
        "room": "205"
    }
    """
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    confirmation_number = str(payload.get("confirmation_number") or "").strip()
    new_room = str(payload.get("room") or "").strip()

    if not confirmation_number:
        return JsonResponse({"error": "confirmation_number is required"}, status=400)
    if not new_room:
        return JsonResponse({"error": "room is required"}, status=400)

    try:
        record = ArrivalRecord.objects.get(confirmation_number=confirmation_number)
    except ArrivalRecord.DoesNotExist:
        return JsonResponse({"error": "Record not found"}, status=404)

    record.room = new_room or None
    record.updated_by = request.user
    record.updated_at = timezone.now()
    record.save()

    return JsonResponse({"success": True, "room": record.room})


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def courtesy_calls(request):
    """
    Page to track courtesy calls for in-house guests.
    """
    today = timezone.localdate()
    context = {
        "section": "guest_experience",
        "subsection": "courtesy_calls",
        "page_title": "Guest Experience - Courtesy Calls",
        "today": today,
    }
    return render(request, "guest_experience/courtesy_calls.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def courtesy_calls_dashboard(request):
    """
    Courtesy Calls dashboard with filters and summary metrics.
    Uses courtesy_calls_api data on the frontend.
    """
    today = timezone.localdate()
    context = {
        "section": "guest_experience",
        "subsection": "courtesy_calls",
        "page_title": "Guest Experience - Courtesy Calls Dashboard",
        "today": today,
    }
    return render(request, "guest_experience/courtesy_calls_dashboard.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def courtesy_comments(request):
    """
    Courtesy call comments page (feedback & issues), filtered on the frontend.
    """
    today = timezone.localdate()
    context = {
        "section": "guest_experience",
        "subsection": "courtesy_calls",
        "page_title": "Guest Experience - Courtesy Call Comments",
        "today": today,
    }
    return render(request, "guest_experience/courtesy_comments.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def in_house_api(request):
    """
    API endpoint for in-house guests.
    Returns guests whose status is 'In-House' for a given date (default: today).
    """
    date_param = request.GET.get("date")
    try:
        if date_param:
            target_date = timezone.datetime.strptime(date_param, "%Y-%m-%d").date()
        else:
            target_date = timezone.localdate()
    except ValueError:
        return JsonResponse({"error": "Invalid date format, expected YYYY-MM-DD."}, status=400)

    # In-house: status In-House (case-insensitive, handles both "In-House" and "in house")
    # Exclude records with "Departed" status
    qs = ArrivalRecord.objects.filter(
        models.Q(status__iexact="in-house") | models.Q(status__iexact="in house")
    ).exclude(
        models.Q(status__iexact="departed")
    )
    qs = qs.filter(
        arrival_date__lte=target_date
    ).filter(
        models.Q(departure_date__isnull=True) | models.Q(departure_date__gte=target_date)
    ).order_by("room")

    data = [
        {
            "room": a.room,
            "guest_name": a.guest_name,
            "confirmation_number": a.confirmation_number,
            "phone": a.phone,
            "country": a.country,
            "travel_agent_name": a.travel_agent_name,
            "arrival_date": a.arrival_date.isoformat() if a.arrival_date else None,
            "departure_date": a.departure_date.isoformat() if a.departure_date else None,
            "nights": a.nights,
            "status": a.status,
            "in_house_since": a.in_house_since.isoformat() if a.in_house_since else None,
            "in_house_by": a.in_house_by.username if a.in_house_by else None,
        }
        for a in qs
    ]
    return JsonResponse({"results": data})


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def departures_api(request):
    """
    API endpoint for departures.
    Returns records with "Departed" status.
    """
    date_param = request.GET.get("date")
    search_query = request.GET.get("search", "").strip()
    
    try:
        if date_param:
            target_date = timezone.datetime.strptime(date_param, "%Y-%m-%d").date()
        else:
            target_date = timezone.localdate()
    except ValueError:
        return JsonResponse({"error": "Invalid date format, expected YYYY-MM-DD."}, status=400)

    # Get records with "Departed" status
    qs = ArrivalRecord.objects.filter(
        models.Q(status__iexact="departed")
    )

    # Filter by departure date if provided
    if date_param:
        qs = qs.filter(departed_at__date=target_date)

    # Apply search filter
    if search_query:
        qs = qs.filter(
            models.Q(room__icontains=search_query) |
            models.Q(guest_name__icontains=search_query) |
            models.Q(confirmation_number__icontains=search_query) |
            models.Q(phone__icontains=search_query) |
            models.Q(country__icontains=search_query) |
            models.Q(travel_agent_name__icontains=search_query) |
            models.Q(departure_method__icontains=search_query)
        )

    qs = qs.order_by("-departed_at", "room")

    data = []
    for a in qs:
        # Format checkout time from departed_at
        checkout_time = ""
        if a.departed_at:
            checkout_time = a.departed_at.strftime("%H:%M")
        
        data.append({
            "room": a.room,
            "guest_name": a.guest_name,
            "confirmation_number": a.confirmation_number,
            "phone": a.phone,
            "email": a.email,
            "country": a.country,
            "nationality": a.nationality,
            "travel_agent_name": a.travel_agent_name,
            "arrival_date": a.arrival_date.isoformat() if a.arrival_date else None,
            "departure_date": a.departure_date.isoformat() if a.departure_date else None,
            "nights": a.nights,
            "status": a.status,
            "departed_at": a.departed_at.isoformat() if a.departed_at else None,
            "checkout_time": checkout_time,
            "departed_by": a.departed_by.username if a.departed_by else None,
            "departure_method": a.departure_method or "",
            "departure_notes": a.departure_notes or "",
            "message_sent_to_guest": a.message_sent_to_guest,
        })

    return JsonResponse({"results": data})


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def courtesy_calls_api(request):
    """
    API endpoint for courtesy call tracking.
    Shows in-house guests and their courtesy call due/completed state.
    """
    now = timezone.now()

    qs = ArrivalRecord.objects.filter(
        models.Q(status__iexact="In-House") | models.Q(status__iexact="in house")
    ).order_by("room")

    data = []
    for a in qs:
        # First courtesy status
        first_status = "Not Scheduled"
        if a.first_courtesy_due_at:
            if a.first_courtesy_done_at:
                first_status = "Completed"
            elif a.first_courtesy_due_at <= now:
                first_status = "Overdue"
            else:
                first_status = "Pending"

        # Second courtesy status
        second_status = "Not Scheduled"
        if a.second_courtesy_due_at:
            if a.second_courtesy_done_at:
                second_status = "Completed"
            elif a.second_courtesy_due_at <= now:
                second_status = "Overdue"
            else:
                second_status = "Pending"

        data.append(
            {
                "room": a.room,
                "guest_name": a.guest_name,
                "confirmation_number": a.confirmation_number,
                "phone": a.phone,
                "country": a.country,
                "arrival_date": a.arrival_date.isoformat() if a.arrival_date else None,
                "departure_date": a.departure_date.isoformat() if a.departure_date else None,
                "nights": a.nights,
                "in_house_since": a.in_house_since.isoformat() if a.in_house_since else None,
                "in_house_by": a.in_house_by.username if a.in_house_by else None,
                "first_courtesy_due_at": a.first_courtesy_due_at.isoformat() if a.first_courtesy_due_at else None,
                "first_courtesy_done_at": a.first_courtesy_done_at.isoformat() if a.first_courtesy_done_at else None,
                "first_courtesy_by": a.first_courtesy_by.username if a.first_courtesy_by else None,
                "first_courtesy_status": first_status,
                "first_courtesy_outcome": a.first_courtesy_outcome,
                "first_courtesy_notes": a.first_courtesy_notes,
                "second_courtesy_due_at": a.second_courtesy_due_at.isoformat() if a.second_courtesy_due_at else None,
                "second_courtesy_done_at": a.second_courtesy_done_at.isoformat() if a.second_courtesy_done_at else None,
                "second_courtesy_by": a.second_courtesy_by.username if a.second_courtesy_by else None,
                "second_courtesy_status": second_status,
                "second_courtesy_outcome": a.second_courtesy_outcome,
                "second_courtesy_notes": a.second_courtesy_notes,
            }
        )

    return JsonResponse({"results": data})


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
@require_POST
def mark_courtesy_done(request):
    """
    Mark a courtesy call as completed.
    Expects JSON:
      {
        "confirmation_number": "...",
        "which": "first"|"second",
        "outcome": "...",
        "notes": "..."
      }
    """
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    confirmation_number = str(payload.get("confirmation_number") or "").strip()
    which = str(payload.get("which") or "").strip().lower()
    outcome = str(payload.get("outcome") or "").strip()
    notes = str(payload.get("notes") or "").strip()
    if not confirmation_number or which not in {"first", "second"}:
        return JsonResponse({"error": "Invalid parameters"}, status=400)

    try:
        record = ArrivalRecord.objects.get(confirmation_number=confirmation_number)
    except ArrivalRecord.DoesNotExist:
        return JsonResponse({"error": "Record not found"}, status=404)

    now = timezone.now()
    if which == "first":
        record.first_courtesy_done_at = now
        record.first_courtesy_by = request.user
        record.first_courtesy_outcome = outcome
        record.first_courtesy_notes = notes
    else:
        record.second_courtesy_done_at = now
        record.second_courtesy_by = request.user
        record.second_courtesy_outcome = outcome
        record.second_courtesy_notes = notes
    record.updated_by = request.user
    record.updated_at = now
    record.save()

    return JsonResponse({"success": True})


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def dashboard_api(request):
    """
    API endpoint for dashboard data including metrics and chart data.
    """
    today = timezone.localdate()
    
    # Get filter parameters
    start_date_str = request.GET.get('start_date', (today - timedelta(days=30)).isoformat())
    end_date_str = request.GET.get('end_date', today.isoformat())
    status_filter = request.GET.get('status', '')
    country_filter = request.GET.get('country', '')
    travel_agent_filter = request.GET.get('travel_agent', '')
    search_query = request.GET.get('search', '')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date = today - timedelta(days=30)
        end_date = today
    
    # Base queryset
    qs = ArrivalRecord.objects.filter(
        arrival_date__gte=start_date,
        arrival_date__lte=end_date
    )
    
    # Apply filters
    if status_filter:
        qs = qs.filter(status__iexact=status_filter)
    
    if country_filter:
        qs = qs.filter(country__icontains=country_filter)
    
    if travel_agent_filter:
        qs = qs.filter(travel_agent_name__icontains=travel_agent_filter)
    
    if search_query:
        qs = qs.filter(
            Q(guest_name__icontains=search_query) |
            Q(confirmation_number__icontains=search_query) |
            Q(room__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(travel_agent_name__icontains=search_query)
        )
    
    # Calculate metrics
    total_arrivals = qs.count()
    
    # Status breakdown
    status_breakdown = qs.values('status').annotate(count=Count('id')).order_by('-count')
    status_data = {item['status'] or 'Unknown': item['count'] for item in status_breakdown}
    
    # Country breakdown (top 10)
    country_breakdown = qs.exclude(country__isnull=True).exclude(country='').values('country').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    country_data = {item['country']: item['count'] for item in country_breakdown}
    
    # Daily arrivals trend
    daily_trend = qs.annotate(date=TruncDate('arrival_date')).values('date').annotate(
        count=Count('id')
    ).order_by('date')
    daily_trend_data = {
        'labels': [item['date'].isoformat() if item['date'] else '' for item in daily_trend],
        'data': [item['count'] for item in daily_trend]
    }
    
    # Nationality breakdown (top 10)
    nationality_breakdown = qs.exclude(nationality__isnull=True).exclude(nationality='').values('nationality').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    nationality_data = {item['nationality']: item['count'] for item in nationality_breakdown}
    
    # Travel agent breakdown (top 10)
    agent_breakdown = qs.exclude(travel_agent_name__isnull=True).exclude(travel_agent_name='').values('travel_agent_name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    agent_data = {item['travel_agent_name']: item['count'] for item in agent_breakdown}
    
    # Expected vs In-House vs Departed
    expected_count = qs.filter(status__iexact='Expected').count()
    in_house_count = qs.filter(
        models.Q(status__iexact='In-House') | models.Q(status__iexact='in house')
    ).count()
    departed_count = qs.filter(status__iexact='Departed').count()
    
    # Average nights
    avg_nights = qs.exclude(nights__isnull=True).aggregate(
        avg_nights=models.Avg('nights')
    )['avg_nights'] or 0
    
    return JsonResponse({
        'metrics': {
            'total_arrivals': total_arrivals,
            'expected': expected_count,
            'in_house': in_house_count,
            'departed': departed_count,
            'avg_nights': round(avg_nights, 1) if avg_nights else 0,
        },
        'charts': {
            'status_breakdown': status_data,
            'country_breakdown': country_data,
            'nationality_breakdown': nationality_data,
            'agent_breakdown': agent_data,
            'daily_trend': daily_trend_data,
        }
    })


# ==================== REPORTS ====================

@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def reports_index(request):
    """Reports landing page."""
    context = {
        "section": "guest_experience",
        "subsection": "reports",
        "page_title": "Guest Experience - Reports",
    }
    return render(request, "guest_experience/reports_index.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def report_arrivals_departures(request):
    """Report 1: Arrivals & Departures Report"""
    today = timezone.localdate()
    
    # Get filters
    start_date_str = request.GET.get('start_date', (today - timedelta(days=7)).isoformat())
    end_date_str = request.GET.get('end_date', (today + timedelta(days=7)).isoformat())
    property_filter = request.GET.get('property', '')
    status_filter = request.GET.get('status', '')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date = today - timedelta(days=7)
        end_date = today + timedelta(days=7)
    
    # Build queryset
    qs = ArrivalRecord.objects.filter(
        Q(arrival_date__gte=start_date, arrival_date__lte=end_date) |
        Q(departure_date__gte=start_date, departure_date__lte=end_date)
    )
    
    if property_filter:
        qs = qs.filter(property_name__icontains=property_filter)
    if status_filter:
        qs = qs.filter(status__iexact=status_filter)
    
    records = qs.order_by('arrival_date', 'room')
    
    # Get unique properties for filter dropdown
    properties = ArrivalRecord.objects.exclude(property_name__isnull=True).exclude(
        property_name=''
    ).values_list('property_name', flat=True).distinct().order_by('property_name')
    
    context = {
        "section": "guest_experience",
        "subsection": "reports",
        "page_title": "Arrivals & Departures Report",
        "records": records,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "property_filter": property_filter,
        "status_filter": status_filter,
        "properties": properties,
    }
    return render(request, "guest_experience/reports/arrivals_departures.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def report_courtesy_call_completion(request):
    """Report 2: Courtesy Call Completion Report"""
    today = timezone.localdate()
    now = timezone.now()
    
    # Get filters
    start_date_str = request.GET.get('start_date', (today - timedelta(days=30)).isoformat())
    end_date_str = request.GET.get('end_date', today.isoformat())
    property_filter = request.GET.get('property', '')
    courtesy_by_filter = request.GET.get('courtesy_by', '')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date = today - timedelta(days=30)
        end_date = today
    
    # Build queryset - guests who arrived in date range
    qs = ArrivalRecord.objects.filter(
        arrival_date__gte=start_date,
        arrival_date__lte=end_date
    )
    
    if property_filter:
        qs = qs.filter(property_name__icontains=property_filter)
    if courtesy_by_filter:
        qs = qs.filter(
            Q(first_courtesy_by__username__icontains=courtesy_by_filter) |
            Q(second_courtesy_by__username__icontains=courtesy_by_filter)
        )
    
    records = []
    for record in qs:
        # Calculate metrics
        first_completed = record.first_courtesy_done_at is not None
        second_completed = record.second_courtesy_done_at is not None
        first_time_taken_minutes = None
        first_time_taken_hours = None
        if first_completed and record.first_courtesy_due_at:
            delta = record.first_courtesy_done_at - record.first_courtesy_due_at
            first_time_taken_minutes = delta.total_seconds() / 60  # minutes
            first_time_taken_hours = first_time_taken_minutes / 60  # hours
        
        records.append({
            'record': record,
            'first_completed': first_completed,
            'second_completed': second_completed,
            'first_time_taken_minutes': first_time_taken_minutes,
            'first_time_taken_hours': first_time_taken_hours,
        })
    
    # Calculate percentages
    total_with_first_due = qs.exclude(first_courtesy_due_at__isnull=True).count()
    first_completed_count = qs.exclude(first_courtesy_done_at__isnull=True).count()
    total_with_second_due = qs.exclude(second_courtesy_due_at__isnull=True).count()
    second_completed_count = qs.exclude(second_courtesy_done_at__isnull=True).count()
    
    first_percentage = (first_completed_count / total_with_first_due * 100) if total_with_first_due > 0 else 0
    second_percentage = (second_completed_count / total_with_second_due * 100) if total_with_second_due > 0 else 0
    
    # Get unique properties and users for filters
    properties = ArrivalRecord.objects.exclude(property_name__isnull=True).exclude(
        property_name=''
    ).values_list('property_name', flat=True).distinct().order_by('property_name')
    
    from django.contrib.auth.models import User
    users = User.objects.filter(
        Q(arrival_records_first_courtesy__isnull=False) |
        Q(arrival_records_second_courtesy__isnull=False)
    ).distinct().order_by('username')
    
    context = {
        "section": "guest_experience",
        "subsection": "reports",
        "page_title": "Courtesy Call Completion Report",
        "records": records,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "property_filter": property_filter,
        "courtesy_by_filter": courtesy_by_filter,
        "properties": properties,
        "users": users,
        "first_percentage": round(first_percentage, 1),
        "second_percentage": round(second_percentage, 1),
        "first_completed_count": first_completed_count,
        "second_completed_count": second_completed_count,
        "total_with_first_due": total_with_first_due,
    }
    return render(request, "guest_experience/reports/courtesy_call_completion.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def report_in_house_guests(request):
    """Report 3: In-House Guests Report"""
    today = timezone.localdate()
    
    # Get filters
    property_filter = request.GET.get('property', '')
    start_date_str = request.GET.get('in_house_since_start', '')
    end_date_str = request.GET.get('in_house_since_end', '')
    
    # Build queryset - currently in-house
    qs = ArrivalRecord.objects.filter(
        Q(status__iexact='In-House') | Q(status__iexact='in house')
    ).exclude(status__iexact='Departed')
    
    if property_filter:
        qs = qs.filter(property_name__icontains=property_filter)
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            qs = qs.filter(in_house_since__date__gte=start_date)
        except (ValueError, TypeError):
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            qs = qs.filter(in_house_since__date__lte=end_date)
        except (ValueError, TypeError):
            pass
    
    records = qs.order_by('room')
    
    # Calculate nights in-house
    for record in records:
        if record.in_house_since:
            delta = timezone.now() - record.in_house_since
            record.nights_in_house = delta.days
        else:
            record.nights_in_house = None
    
    properties = ArrivalRecord.objects.exclude(property_name__isnull=True).exclude(
        property_name=''
    ).values_list('property_name', flat=True).distinct().order_by('property_name')
    
    context = {
        "section": "guest_experience",
        "subsection": "reports",
        "page_title": "In-House Guests Report",
        "records": records,
        "property_filter": property_filter,
        "in_house_since_start": start_date_str,
        "in_house_since_end": end_date_str,
        "properties": properties,
    }
    return render(request, "guest_experience/reports/in_house_guests.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def report_departure_outcomes(request):
    """Report 4: Departure Outcomes Report"""
    today = timezone.localdate()
    
    # Get filters
    start_date_str = request.GET.get('start_date', (today - timedelta(days=30)).isoformat())
    end_date_str = request.GET.get('end_date', today.isoformat())
    property_filter = request.GET.get('property', '')
    departure_method_filter = request.GET.get('departure_method', '')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date = today - timedelta(days=30)
        end_date = today
    
    # Build queryset - departed guests
    qs = ArrivalRecord.objects.filter(
        status__iexact='Departed',
        departed_at__date__gte=start_date,
        departed_at__date__lte=end_date
    )
    
    if property_filter:
        qs = qs.filter(property_name__icontains=property_filter)
    if departure_method_filter:
        qs = qs.filter(departure_method__icontains=departure_method_filter)
    
    records = qs.order_by('-departed_at')
    
    properties = ArrivalRecord.objects.exclude(property_name__isnull=True).exclude(
        property_name=''
    ).values_list('property_name', flat=True).distinct().order_by('property_name')
    
    departure_methods = ArrivalRecord.objects.exclude(
        departure_method__isnull=True
    ).exclude(departure_method='').values_list(
        'departure_method', flat=True
    ).distinct().order_by('departure_method')
    
    context = {
        "section": "guest_experience",
        "subsection": "reports",
        "page_title": "Departure Outcomes Report",
        "records": records,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "property_filter": property_filter,
        "departure_method_filter": departure_method_filter,
        "properties": properties,
        "departure_methods": departure_methods,
    }
    return render(request, "guest_experience/reports/departure_outcomes.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def report_agent_performance(request):
    """Report 5: Agent Performance Report"""
    from django.contrib.auth.models import User
    
    # Get all users who have handled records
    users = User.objects.filter(
        Q(arrival_records_created__isnull=False) |
        Q(arrival_records_updated__isnull=False) |
        Q(arrival_records_in_house__isnull=False) |
        Q(arrival_records_first_courtesy__isnull=False) |
        Q(arrival_records_second_courtesy__isnull=False) |
        Q(arrival_records_departed__isnull=False)
    ).distinct().order_by('username')
    
    performance_data = []
    for user in users:
        created_count = ArrivalRecord.objects.filter(created_by=user).count()
        updated_count = ArrivalRecord.objects.filter(updated_by=user).count()
        in_house_count = ArrivalRecord.objects.filter(in_house_by=user).count()
        first_courtesy_count = ArrivalRecord.objects.filter(first_courtesy_by=user).count()
        second_courtesy_count = ArrivalRecord.objects.filter(second_courtesy_by=user).count()
        departed_count = ArrivalRecord.objects.filter(departed_by=user).count()
        
        total_actions = created_count + updated_count + in_house_count + first_courtesy_count + second_courtesy_count + departed_count
        
        performance_data.append({
            'user': user,
            'created_count': created_count,
            'updated_count': updated_count,
            'in_house_count': in_house_count,
            'first_courtesy_count': first_courtesy_count,
            'second_courtesy_count': second_courtesy_count,
            'departed_count': departed_count,
            'total_actions': total_actions,
        })
    
    # Sort by total actions descending
    performance_data.sort(key=lambda x: x['total_actions'], reverse=True)
    
    context = {
        "section": "guest_experience",
        "subsection": "reports",
        "page_title": "Agent Performance Report",
        "performance_data": performance_data,
    }
    return render(request, "guest_experience/reports/agent_performance.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def report_overdue_actions(request):
    """Report 6: Overdue Actions & Service Lapses"""
    now = timezone.now()
    
    # Get overdue first courtesy calls
    overdue_first = ArrivalRecord.objects.filter(
        first_courtesy_due_at__lt=now,
        first_courtesy_done_at__isnull=True
    ).exclude(status__iexact='Departed')
    
    # Get overdue second courtesy calls
    overdue_second = ArrivalRecord.objects.filter(
        second_courtesy_due_at__lt=now,
        second_courtesy_done_at__isnull=True
    ).exclude(status__iexact='Departed')
    
    # Get overdue departures (departure date passed but not marked as departed)
    overdue_departures = ArrivalRecord.objects.filter(
        departure_date__lt=now.date(),
        status__iexact='In-House'
    )
    
    # Calculate how overdue
    overdue_first_list = []
    for record in overdue_first:
        if record.first_courtesy_due_at:
            delta = now - record.first_courtesy_due_at
            overdue_first_list.append({
                'record': record,
                'overdue_minutes': delta.total_seconds() / 60,
                'overdue_hours': delta.total_seconds() / 3600,
                'overdue_days': delta.days,
            })
    
    overdue_second_list = []
    for record in overdue_second:
        if record.second_courtesy_due_at:
            delta = now - record.second_courtesy_due_at
            overdue_second_list.append({
                'record': record,
                'overdue_minutes': delta.total_seconds() / 60,
                'overdue_hours': delta.total_seconds() / 3600,
                'overdue_days': delta.days,
            })
    
    overdue_departures_list = []
    for record in overdue_departures:
        if record.departure_date:
            delta = now.date() - record.departure_date
            overdue_departures_list.append({
                'record': record,
                'overdue_days': delta.days,
            })
    
    context = {
        "section": "guest_experience",
        "subsection": "reports",
        "page_title": "Overdue Actions & Service Lapses",
        "overdue_first": overdue_first_list,
        "overdue_second": overdue_second_list,
        "overdue_departures": overdue_departures_list,
    }
    return render(request, "guest_experience/reports/overdue_actions.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def report_guest_feedback(request):
    """Report 7: Guest Feedback/Notes Analysis"""
    # Get all records with notes
    records_with_first_notes = ArrivalRecord.objects.exclude(
        first_courtesy_notes__isnull=True
    ).exclude(first_courtesy_notes='')
    
    records_with_second_notes = ArrivalRecord.objects.exclude(
        second_courtesy_notes__isnull=True
    ).exclude(second_courtesy_notes='')
    
    records_with_departure_notes = ArrivalRecord.objects.exclude(
        departure_notes__isnull=True
    ).exclude(departure_notes='')
    
    # Combine all notes for analysis
    all_notes = []
    for record in records_with_first_notes:
        if record.first_courtesy_notes:
            all_notes.append({
                'record': record,
                'note_type': 'First Courtesy',
                'notes': record.first_courtesy_notes,
                'outcome': record.first_courtesy_outcome,
                'date': record.first_courtesy_done_at,
            })
    
    for record in records_with_second_notes:
        if record.second_courtesy_notes:
            all_notes.append({
                'record': record,
                'note_type': 'Second Courtesy',
                'notes': record.second_courtesy_notes,
                'outcome': record.second_courtesy_outcome,
                'date': record.second_courtesy_done_at,
            })
    
    for record in records_with_departure_notes:
        if record.departure_notes:
            all_notes.append({
                'record': record,
                'note_type': 'Departure',
                'notes': record.departure_notes,
                'outcome': None,
                'date': record.departed_at,
            })
    
    # Sort by date descending
    all_notes.sort(key=lambda x: x['date'] if x['date'] else timezone.now(), reverse=True)
    
    context = {
        "section": "guest_experience",
        "subsection": "reports",
        "page_title": "Guest Feedback/Notes Analysis",
        "all_notes": all_notes,
        "first_notes_count": records_with_first_notes.count(),
        "second_notes_count": records_with_second_notes.count(),
        "departure_notes_count": records_with_departure_notes.count(),
    }
    return render(request, "guest_experience/reports/guest_feedback.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def report_nationality_country_breakdown(request):
    """Report 8: Nationality, Country & Market Source Breakdown"""
    # Get filters
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    qs = ArrivalRecord.objects.all()
    
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            qs = qs.filter(arrival_date__gte=start_date)
        except (ValueError, TypeError):
            pass
    
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            qs = qs.filter(arrival_date__lte=end_date)
        except (ValueError, TypeError):
            pass
    
    # Nationality breakdown
    nationality_breakdown = qs.exclude(nationality__isnull=True).exclude(
        nationality=''
    ).values('nationality').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Country breakdown
    country_breakdown = qs.exclude(country__isnull=True).exclude(
        country=''
    ).values('country').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Travel agent breakdown
    agent_breakdown = qs.exclude(travel_agent_name__isnull=True).exclude(
        travel_agent_name=''
    ).values('travel_agent_name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    context = {
        "section": "guest_experience",
        "subsection": "reports",
        "page_title": "Nationality, Country & Market Source Breakdown",
        "nationality_breakdown": nationality_breakdown,
        "country_breakdown": country_breakdown,
        "agent_breakdown": agent_breakdown,
        "start_date": start_date_str,
        "end_date": end_date_str,
    }
    return render(request, "guest_experience/reports/nationality_country_breakdown.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def report_length_of_stay(request):
    """Report 9: Length of Stay Analysis"""
    # Get filters
    property_filter = request.GET.get('property', '')
    nationality_filter = request.GET.get('nationality', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    qs = ArrivalRecord.objects.exclude(nights__isnull=True)
    
    if property_filter:
        qs = qs.filter(property_name__icontains=property_filter)
    if nationality_filter:
        qs = qs.filter(nationality__icontains=nationality_filter)
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            qs = qs.filter(arrival_date__gte=start_date)
        except (ValueError, TypeError):
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            qs = qs.filter(arrival_date__lte=end_date)
        except (ValueError, TypeError):
            pass
    
    # Calculate statistics
    from django.db.models import Avg, Min, Max, Count
    
    overall_stats = qs.aggregate(
        avg_nights=Avg('nights'),
        min_nights=Min('nights'),
        max_nights=Max('nights'),
        total_guests=Count('id')
    )
    
    # By property
    by_property = qs.exclude(property_name__isnull=True).exclude(
        property_name=''
    ).values('property_name').annotate(
        avg_nights=Avg('nights'),
        min_nights=Min('nights'),
        max_nights=Max('nights'),
        count=Count('id')
    ).order_by('-count')
    
    # By nationality
    by_nationality = qs.exclude(nationality__isnull=True).exclude(
        nationality=''
    ).values('nationality').annotate(
        avg_nights=Avg('nights'),
        min_nights=Min('nights'),
        max_nights=Max('nights'),
        count=Count('id')
    ).order_by('-count')
    
    # Distribution (1 night, 2 nights, etc.)
    distribution_raw = qs.values('nights').annotate(
        count=Count('id')
    ).order_by('nights')
    
    # Calculate percentages for distribution
    total_guests = overall_stats['total_guests'] or 1
    distribution = []
    for item in distribution_raw:
        percentage = (item['count'] / total_guests * 100) if total_guests > 0 else 0
        distribution.append({
            'nights': item['nights'],
            'count': item['count'],
            'percentage': round(percentage, 1),
        })
    
    properties = ArrivalRecord.objects.exclude(property_name__isnull=True).exclude(
        property_name=''
    ).values_list('property_name', flat=True).distinct().order_by('property_name')
    
    nationalities = ArrivalRecord.objects.exclude(nationality__isnull=True).exclude(
        nationality=''
    ).values_list('nationality', flat=True).distinct().order_by('nationality')
    
    context = {
        "section": "guest_experience",
        "subsection": "reports",
        "page_title": "Length of Stay Analysis",
        "overall_stats": overall_stats,
        "by_property": by_property,
        "by_nationality": by_nationality,
        "distribution": distribution,
        "property_filter": property_filter,
        "nationality_filter": nationality_filter,
        "start_date": start_date_str,
        "end_date": end_date_str,
        "properties": properties,
        "nationalities": nationalities,
    }
    return render(request, "guest_experience/reports/length_of_stay.html", context)


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def report_contact_completeness(request):
    """Report 10: Contact Data Completeness"""
    # Get all records
    all_records = ArrivalRecord.objects.all()
    
    # Records with missing phone
    missing_phone = all_records.filter(
        Q(phone__isnull=True) | Q(phone='')
    )
    
    # Records with missing email
    missing_email = all_records.filter(
        Q(email__isnull=True) | Q(email='')
    )
    
    # Records with both missing
    missing_both = all_records.filter(
        (Q(phone__isnull=True) | Q(phone='')) &
        (Q(email__isnull=True) | Q(email=''))
    )
    
    # Records with both present
    has_both = all_records.exclude(
        Q(phone__isnull=True) | Q(phone='')
    ).exclude(
        Q(email__isnull=True) | Q(email='')
    )
    
    # Validate email format (basic check)
    import re
    email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    
    invalid_email = []
    for record in all_records.exclude(email__isnull=True).exclude(email=''):
        if not email_pattern.match(record.email):
            invalid_email.append(record)
    
    total_count = all_records.count()
    missing_phone_count = missing_phone.count()
    missing_email_count = missing_email.count()
    missing_both_count = missing_both.count()
    has_both_count = has_both.count()
    invalid_email_count = len(invalid_email)
    
    context = {
        "section": "guest_experience",
        "subsection": "reports",
        "page_title": "Contact Data Completeness",
        "total_count": total_count,
        "missing_phone": missing_phone.order_by('arrival_date'),
        "missing_email": missing_email.order_by('arrival_date'),
        "missing_both": missing_both.order_by('arrival_date'),
        "has_both": has_both.order_by('arrival_date'),
        "invalid_email": invalid_email,
        "missing_phone_count": missing_phone_count,
        "missing_email_count": missing_email_count,
        "missing_both_count": missing_both_count,
        "has_both_count": has_both_count,
        "invalid_email_count": invalid_email_count,
        "phone_completeness": round((total_count - missing_phone_count) / total_count * 100, 1) if total_count > 0 else 0,
        "email_completeness": round((total_count - missing_email_count) / total_count * 100, 1) if total_count > 0 else 0,
    }
    return render(request, "guest_experience/reports/contact_completeness.html", context)


# ==================== EXPORT FUNCTIONS ====================

def _create_excel_workbook():
    """Helper function to create a styled Excel workbook."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is not installed. Please install it to use Excel export.")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    return wb, ws, header_fill, header_font, border


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def export_arrivals_departures(request):
    """Export Arrivals & Departures Report to Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Excel export requires openpyxl. Please install it.", status=500)
    
    today = timezone.localdate()
    start_date_str = request.GET.get('start_date', (today - timedelta(days=7)).isoformat())
    end_date_str = request.GET.get('end_date', (today + timedelta(days=7)).isoformat())
    property_filter = request.GET.get('property', '')
    status_filter = request.GET.get('status', '')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date = today - timedelta(days=7)
        end_date = today + timedelta(days=7)
    
    qs = ArrivalRecord.objects.filter(
        Q(arrival_date__gte=start_date, arrival_date__lte=end_date) |
        Q(departure_date__gte=start_date, departure_date__lte=end_date)
    )
    
    if property_filter:
        qs = qs.filter(property_name__icontains=property_filter)
    if status_filter:
        qs = qs.filter(status__iexact=status_filter)
    
    records = qs.order_by('arrival_date', 'room')
    
    wb, ws, header_fill, header_font, border = _create_excel_workbook()
    ws.title = "Arrivals & Departures"
    
    # Headers
    headers = ['Property', 'Room', 'Guest Name', 'Arrival Date', 'Departure Date', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    for row_idx, record in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=record.property_name or '')
        ws.cell(row=row_idx, column=2, value=record.room or '')
        ws.cell(row=row_idx, column=3, value=record.guest_name or '')
        ws.cell(row=row_idx, column=4, value=record.arrival_date.strftime('%Y-%m-%d') if record.arrival_date else '')
        ws.cell(row=row_idx, column=5, value=record.departure_date.strftime('%Y-%m-%d') if record.departure_date else '')
        ws.cell(row=row_idx, column=6, value=record.status or '')
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="arrivals_departures_{start_date}_{end_date}.xlsx"'
    return response


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def export_courtesy_call_completion(request):
    """Export Courtesy Call Completion Report to Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Excel export requires openpyxl. Please install it.", status=500)
    
    today = timezone.localdate()
    now = timezone.now()
    start_date_str = request.GET.get('start_date', (today - timedelta(days=30)).isoformat())
    end_date_str = request.GET.get('end_date', today.isoformat())
    property_filter = request.GET.get('property', '')
    courtesy_by_filter = request.GET.get('courtesy_by', '')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date = today - timedelta(days=30)
        end_date = today
    
    qs = ArrivalRecord.objects.filter(
        arrival_date__gte=start_date,
        arrival_date__lte=end_date
    )
    
    if property_filter:
        qs = qs.filter(property_name__icontains=property_filter)
    if courtesy_by_filter:
        qs = qs.filter(
            Q(first_courtesy_by__username__icontains=courtesy_by_filter) |
            Q(second_courtesy_by__username__icontains=courtesy_by_filter)
        )
    
    wb, ws, header_fill, header_font, border = _create_excel_workbook()
    ws.title = "Courtesy Call Completion"
    
    headers = ['Room', 'Guest Name', 'First Call Status', 'First Call Time (min)', 'First Call Time (hrs)', 
               'First Outcome', 'Second Call Status', 'Second Outcome', 'First Notes', 'Second Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    for record in qs:
        first_completed = record.first_courtesy_done_at is not None
        second_completed = record.second_courtesy_done_at is not None
        first_time_minutes = None
        first_time_hours = None
        if first_completed and record.first_courtesy_due_at:
            delta = record.first_courtesy_done_at - record.first_courtesy_due_at
            first_time_minutes = delta.total_seconds() / 60
            first_time_hours = first_time_minutes / 60
        
        first_status = "Completed" if first_completed else ("Pending" if record.first_courtesy_due_at else "Not Scheduled")
        second_status = "Completed" if second_completed else ("Pending" if record.second_courtesy_due_at else "Not Scheduled")
        
        ws.cell(row=row_idx, column=1, value=record.room or '')
        ws.cell(row=row_idx, column=2, value=record.guest_name or '')
        ws.cell(row=row_idx, column=3, value=first_status)
        ws.cell(row=row_idx, column=4, value=round(first_time_minutes, 1) if first_time_minutes else '')
        ws.cell(row=row_idx, column=5, value=round(first_time_hours, 1) if first_time_hours else '')
        ws.cell(row=row_idx, column=6, value=record.first_courtesy_outcome or '')
        ws.cell(row=row_idx, column=7, value=second_status)
        ws.cell(row=row_idx, column=8, value=record.second_courtesy_outcome or '')
        ws.cell(row=row_idx, column=9, value=record.first_courtesy_notes or '')
        ws.cell(row=row_idx, column=10, value=record.second_courtesy_notes or '')
        row_idx += 1
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="courtesy_call_completion_{start_date}_{end_date}.xlsx"'
    return response


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def export_in_house_guests(request):
    """Export In-House Guests Report to Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Excel export requires openpyxl. Please install it.", status=500)
    
    property_filter = request.GET.get('property', '')
    start_date_str = request.GET.get('in_house_since_start', '')
    end_date_str = request.GET.get('in_house_since_end', '')
    
    qs = ArrivalRecord.objects.filter(
        Q(status__iexact='In-House') | Q(status__iexact='in house')
    ).exclude(status__iexact='Departed')
    
    if property_filter:
        qs = qs.filter(property_name__icontains=property_filter)
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            qs = qs.filter(in_house_since__date__gte=start_date)
        except (ValueError, TypeError):
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            qs = qs.filter(in_house_since__date__lte=end_date)
        except (ValueError, TypeError):
            pass
    
    records = qs.order_by('room')
    
    wb, ws, header_fill, header_font, border = _create_excel_workbook()
    ws.title = "In-House Guests"
    
    headers = ['Room', 'Guest Name', 'In-House Since', 'Nights In-House', 'Total Nights', 'Status']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    for record in records:
        nights_in_house = None
        if record.in_house_since:
            delta = timezone.now() - record.in_house_since
            nights_in_house = delta.days
        
        ws.cell(row=row_idx, column=1, value=record.room or '')
        ws.cell(row=row_idx, column=2, value=record.guest_name or '')
        ws.cell(row=row_idx, column=3, value=record.in_house_since.strftime('%Y-%m-%d %H:%M') if record.in_house_since else '')
        ws.cell(row=row_idx, column=4, value=nights_in_house if nights_in_house is not None else '')
        ws.cell(row=row_idx, column=5, value=record.nights if record.nights else '')
        ws.cell(row=row_idx, column=6, value=record.status or '')
        row_idx += 1
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="in_house_guests.xlsx"'
    return response


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def export_departure_outcomes(request):
    """Export Departure Outcomes Report to Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Excel export requires openpyxl. Please install it.", status=500)
    
    today = timezone.localdate()
    start_date_str = request.GET.get('start_date', (today - timedelta(days=30)).isoformat())
    end_date_str = request.GET.get('end_date', today.isoformat())
    property_filter = request.GET.get('property', '')
    departure_method_filter = request.GET.get('departure_method', '')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date = today - timedelta(days=30)
        end_date = today
    
    qs = ArrivalRecord.objects.filter(
        status__iexact='Departed',
        departed_at__date__gte=start_date,
        departed_at__date__lte=end_date
    )
    
    if property_filter:
        qs = qs.filter(property_name__icontains=property_filter)
    if departure_method_filter:
        qs = qs.filter(departure_method__icontains=departure_method_filter)
    
    records = qs.order_by('-departed_at')
    
    wb, ws, header_fill, header_font, border = _create_excel_workbook()
    ws.title = "Departure Outcomes"
    
    headers = ['Departed At', 'Departed By', 'Room', 'Guest Name', 'Departure Method', 'Message Sent', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    for record in records:
        ws.cell(row=row_idx, column=1, value=record.departed_at.strftime('%Y-%m-%d %H:%M') if record.departed_at else '')
        ws.cell(row=row_idx, column=2, value=record.departed_by.username if record.departed_by else '')
        ws.cell(row=row_idx, column=3, value=record.room or '')
        ws.cell(row=row_idx, column=4, value=record.guest_name or '')
        ws.cell(row=row_idx, column=5, value=record.departure_method or '')
        ws.cell(row=row_idx, column=6, value='Yes' if record.message_sent_to_guest else 'No')
        ws.cell(row=row_idx, column=7, value=record.departure_notes or '')
        row_idx += 1
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="departure_outcomes_{start_date}_{end_date}.xlsx"'
    return response


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def export_agent_performance(request):
    """Export Agent Performance Report to Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Excel export requires openpyxl. Please install it.", status=500)
    
    from django.contrib.auth.models import User
    
    users = User.objects.filter(
        Q(arrival_records_created__isnull=False) |
        Q(arrival_records_updated__isnull=False) |
        Q(arrival_records_in_house__isnull=False) |
        Q(arrival_records_first_courtesy__isnull=False) |
        Q(arrival_records_second_courtesy__isnull=False) |
        Q(arrival_records_departed__isnull=False)
    ).distinct().order_by('username')
    
    wb, ws, header_fill, header_font, border = _create_excel_workbook()
    ws.title = "Agent Performance"
    
    headers = ['Agent', 'Created', 'Updated', 'Marked In-House', 'First Courtesy', 'Second Courtesy', 'Departed', 'Total Actions']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    for user in users:
        created_count = ArrivalRecord.objects.filter(created_by=user).count()
        updated_count = ArrivalRecord.objects.filter(updated_by=user).count()
        in_house_count = ArrivalRecord.objects.filter(in_house_by=user).count()
        first_courtesy_count = ArrivalRecord.objects.filter(first_courtesy_by=user).count()
        second_courtesy_count = ArrivalRecord.objects.filter(second_courtesy_by=user).count()
        departed_count = ArrivalRecord.objects.filter(departed_by=user).count()
        total_actions = created_count + updated_count + in_house_count + first_courtesy_count + second_courtesy_count + departed_count
        
        ws.cell(row=row_idx, column=1, value=user.username)
        ws.cell(row=row_idx, column=2, value=created_count)
        ws.cell(row=row_idx, column=3, value=updated_count)
        ws.cell(row=row_idx, column=4, value=in_house_count)
        ws.cell(row=row_idx, column=5, value=first_courtesy_count)
        ws.cell(row=row_idx, column=6, value=second_courtesy_count)
        ws.cell(row=row_idx, column=7, value=departed_count)
        ws.cell(row=row_idx, column=8, value=total_actions)
        row_idx += 1
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="agent_performance.xlsx"'
    return response


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def export_overdue_actions(request):
    """Export Overdue Actions Report to Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Excel export requires openpyxl. Please install it.", status=500)
    
    now = timezone.now()
    
    overdue_first = ArrivalRecord.objects.filter(
        first_courtesy_due_at__lt=now,
        first_courtesy_done_at__isnull=True
    ).exclude(status__iexact='Departed')
    
    overdue_second = ArrivalRecord.objects.filter(
        second_courtesy_due_at__lt=now,
        second_courtesy_done_at__isnull=True
    ).exclude(status__iexact='Departed')
    
    overdue_departures = ArrivalRecord.objects.filter(
        departure_date__lt=now.date(),
        status__iexact='In-House'
    )
    
    wb, ws, header_fill, header_font, border = _create_excel_workbook()
    ws.title = "Overdue First Calls"
    
    # First sheet - Overdue First Calls
    headers = ['Room', 'Guest Name', 'Due At', 'Overdue (days)', 'Overdue (hours)', 'Overdue (minutes)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    # Convert queryset to list to ensure evaluation
    overdue_first_list = list(overdue_first)
    for record in overdue_first_list:
        if record.first_courtesy_due_at:
            delta = now - record.first_courtesy_due_at
            ws.cell(row=row_idx, column=1, value=record.room or '')
            ws.cell(row=row_idx, column=2, value=record.guest_name or '')
            ws.cell(row=row_idx, column=3, value=record.first_courtesy_due_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_idx, column=4, value=delta.days)
            ws.cell(row=row_idx, column=5, value=round(delta.total_seconds() / 3600, 1))
            ws.cell(row=row_idx, column=6, value=round(delta.total_seconds() / 60, 1))
            row_idx += 1
    
    # Second sheet - Overdue Second Calls
    ws2 = wb.create_sheet("Overdue Second Calls")
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    # Convert queryset to list to ensure evaluation
    overdue_second_list = list(overdue_second)
    for record in overdue_second_list:
        if record.second_courtesy_due_at:
            delta = now - record.second_courtesy_due_at
            ws2.cell(row=row_idx, column=1, value=record.room or '')
            ws2.cell(row=row_idx, column=2, value=record.guest_name or '')
            ws2.cell(row=row_idx, column=3, value=record.second_courtesy_due_at.strftime('%Y-%m-%d %H:%M'))
            ws2.cell(row=row_idx, column=4, value=delta.days)
            ws2.cell(row=row_idx, column=5, value=round(delta.total_seconds() / 3600, 1))
            ws2.cell(row=row_idx, column=6, value=round(delta.total_seconds() / 60, 1))
            row_idx += 1
    
    # Third sheet - Overdue Departures
    ws3 = wb.create_sheet("Overdue Departures")
    headers3 = ['Room', 'Guest Name', 'Departure Date', 'Overdue (days)']
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    # Convert queryset to list to ensure evaluation
    overdue_departures_list = list(overdue_departures)
    for record in overdue_departures_list:
        if record.departure_date:
            delta = now.date() - record.departure_date
            ws3.cell(row=row_idx, column=1, value=record.room or '')
            ws3.cell(row=row_idx, column=2, value=record.guest_name or '')
            ws3.cell(row=row_idx, column=3, value=record.departure_date.strftime('%Y-%m-%d'))
            ws3.cell(row=row_idx, column=4, value=delta.days)
            row_idx += 1
    
    for sheet in [ws, ws2, ws3]:
        for col in range(1, 7):
            sheet.column_dimensions[get_column_letter(col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="overdue_actions.xlsx"'
    return response


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def export_guest_feedback(request):
    """Export Guest Feedback Report to Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Excel export requires openpyxl. Please install it.", status=500)
    
    records_with_first_notes = ArrivalRecord.objects.exclude(
        first_courtesy_notes__isnull=True
    ).exclude(first_courtesy_notes='')
    
    records_with_second_notes = ArrivalRecord.objects.exclude(
        second_courtesy_notes__isnull=True
    ).exclude(second_courtesy_notes='')
    
    records_with_departure_notes = ArrivalRecord.objects.exclude(
        departure_notes__isnull=True
    ).exclude(departure_notes='')
    
    wb, ws, header_fill, header_font, border = _create_excel_workbook()
    ws.title = "Guest Feedback"
    
    headers = ['Date', 'Type', 'Room', 'Guest Name', 'Outcome', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    all_notes = []
    
    for record in records_with_first_notes:
        if record.first_courtesy_notes:
            all_notes.append({
                'date': record.first_courtesy_done_at,
                'type': 'First Courtesy',
                'room': record.room,
                'guest': record.guest_name,
                'outcome': record.first_courtesy_outcome,
                'notes': record.first_courtesy_notes,
            })
    
    for record in records_with_second_notes:
        if record.second_courtesy_notes:
            all_notes.append({
                'date': record.second_courtesy_done_at,
                'type': 'Second Courtesy',
                'room': record.room,
                'guest': record.guest_name,
                'outcome': record.second_courtesy_outcome,
                'notes': record.second_courtesy_notes,
            })
    
    for record in records_with_departure_notes:
        if record.departure_notes:
            all_notes.append({
                'date': record.departed_at,
                'type': 'Departure',
                'room': record.room,
                'guest': record.guest_name,
                'outcome': None,
                'notes': record.departure_notes,
            })
    
    all_notes.sort(key=lambda x: x['date'] if x['date'] else timezone.now(), reverse=True)
    
    for note in all_notes:
        ws.cell(row=row_idx, column=1, value=note['date'].strftime('%Y-%m-%d %H:%M') if note['date'] else '')
        ws.cell(row=row_idx, column=2, value=note['type'])
        ws.cell(row=row_idx, column=3, value=note['room'] or '')
        ws.cell(row=row_idx, column=4, value=note['guest'] or '')
        ws.cell(row=row_idx, column=5, value=note['outcome'] or '')
        ws.cell(row=row_idx, column=6, value=note['notes'])
        row_idx += 1
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="guest_feedback.xlsx"'
    return response


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def export_nationality_country_breakdown(request):
    """Export Nationality, Country & Market Source Breakdown to Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Excel export requires openpyxl. Please install it.", status=500)
    
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    qs = ArrivalRecord.objects.all()
    
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            qs = qs.filter(arrival_date__gte=start_date)
        except (ValueError, TypeError):
            pass
    
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            qs = qs.filter(arrival_date__lte=end_date)
        except (ValueError, TypeError):
            pass
    
    nationality_breakdown = qs.exclude(nationality__isnull=True).exclude(
        nationality=''
    ).values('nationality').annotate(
        count=Count('id')
    ).order_by('-count')
    
    country_breakdown = qs.exclude(country__isnull=True).exclude(
        country=''
    ).values('country').annotate(
        count=Count('id')
    ).order_by('-count')
    
    agent_breakdown = qs.exclude(travel_agent_name__isnull=True).exclude(
        travel_agent_name=''
    ).values('travel_agent_name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    wb, ws, header_fill, header_font, border = _create_excel_workbook()
    ws.title = "Nationality Breakdown"
    
    headers = ['Nationality', 'Count']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    for item in nationality_breakdown:
        ws.cell(row=row_idx, column=1, value=item['nationality'])
        ws.cell(row=row_idx, column=2, value=item['count'])
        row_idx += 1
    
    ws2 = wb.create_sheet("Country Breakdown")
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    for item in country_breakdown:
        ws2.cell(row=row_idx, column=1, value=item['country'])
        ws2.cell(row=row_idx, column=2, value=item['count'])
        row_idx += 1
    
    ws3 = wb.create_sheet("Travel Agent Breakdown")
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    for item in agent_breakdown:
        ws3.cell(row=row_idx, column=1, value=item['travel_agent_name'])
        ws3.cell(row=row_idx, column=2, value=item['count'])
        row_idx += 1
    
    for sheet in [ws, ws2, ws3]:
        for col in range(1, 3):
            sheet.column_dimensions[get_column_letter(col)].width = 30
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="nationality_country_breakdown.xlsx"'
    return response


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def export_length_of_stay(request):
    """Export Length of Stay Analysis to Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Excel export requires openpyxl. Please install it.", status=500)
    
    property_filter = request.GET.get('property', '')
    nationality_filter = request.GET.get('nationality', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    qs = ArrivalRecord.objects.exclude(nights__isnull=True)
    
    if property_filter:
        qs = qs.filter(property_name__icontains=property_filter)
    if nationality_filter:
        qs = qs.filter(nationality__icontains=nationality_filter)
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            qs = qs.filter(arrival_date__gte=start_date)
        except (ValueError, TypeError):
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            qs = qs.filter(arrival_date__lte=end_date)
        except (ValueError, TypeError):
            pass
    
    overall_stats = qs.aggregate(
        avg_nights=Avg('nights'),
        min_nights=Min('nights'),
        max_nights=Max('nights'),
        total_guests=Count('id')
    )
    
    by_property = qs.exclude(property_name__isnull=True).exclude(
        property_name=''
    ).values('property_name').annotate(
        avg_nights=Avg('nights'),
        min_nights=Min('nights'),
        max_nights=Max('nights'),
        count=Count('id')
    ).order_by('-count')
    
    by_nationality = qs.exclude(nationality__isnull=True).exclude(
        nationality=''
    ).values('nationality').annotate(
        avg_nights=Avg('nights'),
        min_nights=Min('nights'),
        max_nights=Max('nights'),
        count=Count('id')
    ).order_by('-count')
    
    distribution_raw = qs.values('nights').annotate(
        count=Count('id')
    ).order_by('nights')
    
    total_guests = overall_stats['total_guests'] or 1
    distribution = []
    for item in distribution_raw:
        percentage = (item['count'] / total_guests * 100) if total_guests > 0 else 0
        distribution.append({
            'nights': item['nights'],
            'count': item['count'],
            'percentage': round(percentage, 1),
        })
    
    wb, ws, header_fill, header_font, border = _create_excel_workbook()
    ws.title = "Overall Statistics"
    
    ws.cell(row=1, column=1, value="Metric")
    ws.cell(row=1, column=2, value="Value")
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).font = header_font
    
    ws.cell(row=2, column=1, value="Average Nights")
    ws.cell(row=2, column=2, value=round(overall_stats['avg_nights'], 1) if overall_stats['avg_nights'] else '')
    ws.cell(row=3, column=1, value="Minimum Nights")
    ws.cell(row=3, column=2, value=overall_stats['min_nights'] if overall_stats['min_nights'] else '')
    ws.cell(row=4, column=1, value="Maximum Nights")
    ws.cell(row=4, column=2, value=overall_stats['max_nights'] if overall_stats['max_nights'] else '')
    ws.cell(row=5, column=1, value="Total Guests")
    ws.cell(row=5, column=2, value=overall_stats['total_guests'] or 0)
    
    ws2 = wb.create_sheet("By Property")
    headers = ['Property', 'Avg Nights', 'Min Nights', 'Max Nights', 'Count']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    for item in by_property:
        ws2.cell(row=row_idx, column=1, value=item['property_name'])
        ws2.cell(row=row_idx, column=2, value=round(item['avg_nights'], 1) if item['avg_nights'] else '')
        ws2.cell(row=row_idx, column=3, value=item['min_nights'] if item['min_nights'] else '')
        ws2.cell(row=row_idx, column=4, value=item['max_nights'] if item['max_nights'] else '')
        ws2.cell(row=row_idx, column=5, value=item['count'])
        row_idx += 1
    
    ws3 = wb.create_sheet("By Nationality")
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    for item in by_nationality:
        ws3.cell(row=row_idx, column=1, value=item['nationality'])
        ws3.cell(row=row_idx, column=2, value=round(item['avg_nights'], 1) if item['avg_nights'] else '')
        ws3.cell(row=row_idx, column=3, value=item['min_nights'] if item['min_nights'] else '')
        ws3.cell(row=row_idx, column=4, value=item['max_nights'] if item['max_nights'] else '')
        ws3.cell(row=row_idx, column=5, value=item['count'])
        row_idx += 1
    
    ws4 = wb.create_sheet("Distribution")
    headers_dist = ['Nights', 'Count', 'Percentage']
    for col, header in enumerate(headers_dist, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    for item in distribution:
        ws4.cell(row=row_idx, column=1, value=item['nights'])
        ws4.cell(row=row_idx, column=2, value=item['count'])
        ws4.cell(row=row_idx, column=3, value=item['percentage'])
        row_idx += 1
    
    for sheet in [ws, ws2, ws3, ws4]:
        for col in range(1, 6):
            sheet.column_dimensions[get_column_letter(col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="length_of_stay.xlsx"'
    return response


@login_required
@permission_required("accounts.view_hotel_management", raise_exception=True)
def export_contact_completeness(request):
    """Export Contact Data Completeness Report to Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Excel export requires openpyxl. Please install it.", status=500)
    
    all_records = ArrivalRecord.objects.all()
    missing_phone = all_records.filter(Q(phone__isnull=True) | Q(phone=''))
    missing_email = all_records.filter(Q(email__isnull=True) | Q(email=''))
    missing_both = all_records.filter(
        (Q(phone__isnull=True) | Q(phone='')) &
        (Q(email__isnull=True) | Q(email=''))
    )
    
    import re
    email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    invalid_email = []
    for record in all_records.exclude(email__isnull=True).exclude(email=''):
        if not email_pattern.match(record.email):
            invalid_email.append(record)
    
    wb, ws, header_fill, header_font, border = _create_excel_workbook()
    ws.title = "Missing Phone"
    
    headers = ['Room', 'Guest Name', 'Arrival Date', 'Email']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    # Convert queryset to list to ensure evaluation
    missing_phone_list = list(missing_phone.order_by('arrival_date'))
    for record in missing_phone_list:
        ws.cell(row=row_idx, column=1, value=record.room or '')
        ws.cell(row=row_idx, column=2, value=record.guest_name or '')
        ws.cell(row=row_idx, column=3, value=record.arrival_date.strftime('%Y-%m-%d') if record.arrival_date else '')
        ws.cell(row=row_idx, column=4, value=record.email or '')
        row_idx += 1
    
    ws2 = wb.create_sheet("Missing Email")
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    # Convert queryset to list to ensure evaluation
    missing_email_list = list(missing_email.order_by('arrival_date'))
    for record in missing_email_list:
        ws2.cell(row=row_idx, column=1, value=record.room or '')
        ws2.cell(row=row_idx, column=2, value=record.guest_name or '')
        ws2.cell(row=row_idx, column=3, value=record.arrival_date.strftime('%Y-%m-%d') if record.arrival_date else '')
        ws2.cell(row=row_idx, column=4, value=record.phone or '')
        row_idx += 1
    
    ws3 = wb.create_sheet("Missing Both")
    headers3 = ['Room', 'Guest Name', 'Arrival Date', 'Confirmation #']
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    # Convert queryset to list to ensure evaluation
    missing_both_list = list(missing_both.order_by('arrival_date'))
    for record in missing_both_list:
        ws3.cell(row=row_idx, column=1, value=record.room or '')
        ws3.cell(row=row_idx, column=2, value=record.guest_name or '')
        ws3.cell(row=row_idx, column=3, value=record.arrival_date.strftime('%Y-%m-%d') if record.arrival_date else '')
        ws3.cell(row=row_idx, column=4, value=record.confirmation_number or '')
        row_idx += 1
    
    ws4 = wb.create_sheet("Invalid Email")
    headers4 = ['Room', 'Guest Name', 'Invalid Email', 'Phone']
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    # invalid_email is already a list, so no need to convert
    for record in invalid_email:
        ws4.cell(row=row_idx, column=1, value=record.room or '')
        ws4.cell(row=row_idx, column=2, value=record.guest_name or '')
        ws4.cell(row=row_idx, column=3, value=record.email)
        ws4.cell(row=row_idx, column=4, value=record.phone or '')
        row_idx += 1
    
    for sheet in [ws, ws2, ws3, ws4]:
        for col in range(1, 5):
            sheet.column_dimensions[get_column_letter(col)].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="contact_completeness.xlsx"'
    return response
