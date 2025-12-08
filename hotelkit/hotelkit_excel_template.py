"""
Generate an Excel template for Hotelkit guest requests and repair tracking.
The workbook includes data entry tables and two dashboards with summary
metrics and charts. Install dependencies then run:

    pip install openpyxl
    python hotelkit_excel_template.py
"""

import io
from datetime import date, timedelta
from typing import Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(cells: Iterable) -> None:
    for cell in cells:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def auto_fit(ws) -> None:
    for column_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in column_cells)
        adjusted = max(10, min(max_len + 2, 40))
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted


def add_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def populate_guest_requests(ws) -> None:
    headers = [
        "Request ID", "Guest Name", "Room", "Request Type",
        "Priority", "Status", "Created Date", "Due Date",
        "Completed Date", "Notes",
    ]
    ws.append(headers)
    style_header(ws[1])

    today = date.today()
    samples: List[Tuple] = [
        ("GR-001", "Anna Li", "1205", "Housekeeping", "High", "Open",
         today - timedelta(days=1), today + timedelta(days=1), None, "Extra towels"),
        ("GR-002", "Mohamed Ali", "803", "Maintenance", "Medium", "In Progress",
         today - timedelta(days=2), today + timedelta(days=2), None, "AC warm"),
        ("GR-003", "Sara Kim", "402", "Concierge", "Low", "Closed",
         today - timedelta(days=5), today - timedelta(days=1), today - timedelta(days=1),
         "Restaurant booking done"),
    ]
    for row in samples:
        ws.append(row)

    add_table(ws, "GuestRequests", f"A1:J{len(samples)+1}")
    auto_fit(ws)


def populate_repairs(ws) -> None:
    headers = [
        "Ticket ID", "Area", "Asset", "Issue", "Priority",
        "Status", "Reported Date", "Target Date", "Completed Date",
        "Technician", "Notes",
    ]
    ws.append(headers)
    style_header(ws[1])

    today = date.today()
    samples: List[Tuple] = [
        ("RP-1001", "Lobby", "Elevator A", "Door sensor fault", "High",
         "In Progress", today - timedelta(days=1), today + timedelta(days=2), None,
         "J. Ortega", "Waiting on part"),
        ("RP-1002", "Pool", "Pump 2", "Low pressure", "High",
         "Open", today, today + timedelta(days=3), None, "K. Singh", None),
        ("RP-1003", "Rooms", "HVAC", "No cooling", "Medium",
         "Closed", today - timedelta(days=4), today - timedelta(days=1),
         today - timedelta(days=1), "L. Perez", "Recharged gas"),
    ]
    for row in samples:
        ws.append(row)

    add_table(ws, "Repairs", f"A1:K{len(samples)+1}")
    auto_fit(ws)


def populate_lookups(ws) -> None:
    ws.append(["Status Options", "Priority Options", "Request Types"])
    style_header(ws[1])
    rows = [
        ("Open", "Low", "Housekeeping"),
        ("In Progress", "Medium", "Maintenance"),
        ("Closed", "High", "Concierge"),
        ("On Hold", "Critical", "IT"),
    ]
    for row in rows:
        ws.append(row)
    auto_fit(ws)


def summary_block(ws, start_row: int, title: str, data_ref: str) -> int:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)

    labels = ["Open", "In Progress", "Closed", "On Hold"]
    for i, label in enumerate(labels, start=1):
        ws.cell(row=start_row + i, column=1, value=label)
        formula = f'=COUNTIFS({data_ref},"{label}")'
        ws.cell(row=start_row + i, column=2, value=formula)

    ws.cell(row=start_row + len(labels) + 2, column=1, value="Total")
    ws.cell(row=start_row + len(labels) + 2, column=2, value=f"=SUM(B{start_row+1}:B{start_row+len(labels)})")

    style_header([ws.cell(row=start_row, column=1)])
    return start_row + len(labels) + 4


def add_guest_dashboard(wb) -> None:
    ws = wb.create_sheet("Guest Dashboard")
    next_row = summary_block(ws, 1, "Guest Requests by Status", "GuestRequests[Status]")

    ws.cell(row=next_row, column=1, value="Requests by Type")
    style_header([ws.cell(row=next_row, column=1)])
    type_labels = ["Housekeeping", "Maintenance", "Concierge", "IT"]
    for i, label in enumerate(type_labels, start=1):
        ws.cell(row=next_row + i, column=1, value=label)
        ws.cell(row=next_row + i, column=2,
                value=f'=COUNTIF(GuestRequests[Request Type],"{label}")')

    # Status chart
    status_chart = BarChart()
    status_chart.title = "Status Distribution"
    data = Reference(ws, min_col=2, min_row=2, max_row=5)
    cats = Reference(ws, min_col=1, min_row=2, max_row=5)
    status_chart.add_data(data, titles_from_data=False)
    status_chart.set_categories(cats)
    status_chart.height = 7
    status_chart.width = 12
    ws.add_chart(status_chart, "D2")

    # Requests over time
    ws.cell(row=20, column=1, value="Requests by Date")
    style_header([ws.cell(row=20, column=1)])
    ws.cell(row=21, column=1, value="Date")
    ws.cell(row=21, column=2, value="Count")
    style_header(ws[21])
    for i in range(7):
        ws.cell(row=22 + i, column=1, value=date.today() - timedelta(days=6 - i))
        ws.cell(row=22 + i, column=2,
                value=f'=COUNTIF(GuestRequests[Created Date],A{22 + i})')

    line_chart = LineChart()
    line_chart.title = "Requests Trend (Last 7 days)"
    data = Reference(ws, min_col=2, min_row=21, max_row=28)
    cats = Reference(ws, min_col=1, min_row=22, max_row=28)
    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(cats)
    line_chart.height = 7
    line_chart.width = 12
    ws.add_chart(line_chart, "D12")

    auto_fit(ws)


def add_repair_dashboard(wb) -> None:
    ws = wb.create_sheet("Repairs Dashboard")
    next_row = summary_block(ws, 1, "Repairs by Status", "Repairs[Status]")

    ws.cell(row=next_row, column=1, value="Repairs by Priority")
    style_header([ws.cell(row=next_row, column=1)])
    priorities = ["Low", "Medium", "High", "Critical"]
    for i, label in enumerate(priorities, start=1):
        ws.cell(row=next_row + i, column=1, value=label)
        ws.cell(row=next_row + i, column=2,
                value=f'=COUNTIF(Repairs[Priority],"{label}")')

    status_chart = BarChart()
    status_chart.title = "Status Distribution"
    data = Reference(ws, min_col=2, min_row=2, max_row=5)
    cats = Reference(ws, min_col=1, min_row=2, max_row=5)
    status_chart.add_data(data, titles_from_data=False)
    status_chart.set_categories(cats)
    status_chart.height = 7
    status_chart.width = 12
    ws.add_chart(status_chart, "D2")

    ws.cell(row=20, column=1, value="Repairs by Area")
    style_header([ws.cell(row=20, column=1)])
    ws.cell(row=21, column=1, value="Area")
    ws.cell(row=21, column=2, value="Count")
    style_header(ws[21])
    areas = ["Lobby", "Pool", "Rooms", "Back of House"]
    for i, area in enumerate(areas, start=1):
        ws.cell(row=21 + i, column=1, value=area)
        ws.cell(row=21 + i, column=2,
                value=f'=COUNTIF(Repairs[Area],"{area}")')

    bar_chart = BarChart()
    bar_chart.title = "Repairs by Area"
    data = Reference(ws, min_col=2, min_row=21, max_row=25)
    cats = Reference(ws, min_col=1, min_row=22, max_row=25)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    bar_chart.height = 7
    bar_chart.width = 12
    ws.add_chart(bar_chart, "D12")

    auto_fit(ws)


def build_workbook() -> Workbook:
    """Create the hotelkit template workbook with data-entry sheets and dashboards."""
    wb = Workbook()
    wb.remove(wb.active)
    populate_guest_requests(wb.create_sheet("Guest Requests"))
    populate_repairs(wb.create_sheet("Repairs"))
    populate_lookups(wb.create_sheet("Lookups"))
    add_guest_dashboard(wb)
    add_repair_dashboard(wb)
    return wb


def render_template_bytes() -> bytes:
    """Return the workbook as Excel bytes (in-memory)."""
    wb = build_workbook()
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def main() -> None:
    buffer = render_template_bytes()
    with open("hotelkit_template.xlsx", "wb") as f:
        f.write(buffer)
    print("Created hotelkit_template.xlsx")


if __name__ == "__main__":
    main()

