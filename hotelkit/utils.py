import pandas as pd
from datetime import datetime
from django.utils import timezone
from .models import RepairRequest


def parse_excel_file(file_path):
    """
    Parse Excel file and return DataFrame with proper column mapping.
    """
    try:
        # Support both file paths and file-like uploads
        filename = getattr(file_path, 'name', str(file_path))
        lower_name = filename.lower()

        # Ensure file-like objects start at the beginning
        if hasattr(file_path, 'seek'):
            try:
                file_path.seek(0)
            except Exception:
                pass

        # Choose engine explicitly to avoid optional dependency ambiguity
        if lower_name.endswith('.xlsx'):
            engine = 'openpyxl'
        elif lower_name.endswith('.xls'):
            engine = 'xlrd'
        else:
            raise Exception('Unsupported file type. Please upload .xlsx or .xls files.')

        # Read Excel file with explicit engine
        df = pd.read_excel(file_path, engine=engine)
        
        # Map column names to model fields
        column_mapping = {
            'Position': 'position',
            'ID': 'id_field',
            'Creator': 'creator',
            'Recipients': 'recipients',
            'Location': 'location',
            'Location path': 'location_path',
            'Type': 'type',
            'Type path': 'type_path',
            'Assets': 'assets',
            'Ticket': 'ticket',
            'Creation date': 'creation_date',
            'Priority': 'priority',
            'State': 'state',
            'Latest state change user': 'latest_state_change_user',
            'Latest state change time': 'latest_state_change_time',
            'Time accepted': 'time_accepted',
            'Time in progress': 'time_in_progress',
            'Time done': 'time_done',
            'Time "in evaluation"': 'time_in_evaluation',
            'Text': 'text',
            'Link': 'link',
            'Submitted result': 'submitted_result',
            'Comments': 'comments',
            'Parking reason': 'parking_reason',
            'Parking information': 'parking_information',
        }
        
        # Rename columns
        df = df.rename(columns=column_mapping)
        
        # Convert datetime columns
        datetime_columns = [
            'creation_date', 'latest_state_change_time', 'time_accepted',
            'time_in_progress', 'time_done', 'time_in_evaluation'
        ]
        
        for col in datetime_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
        
        return df
        
    except ImportError as e:
        # Provide clearer guidance when engine backends are missing
        missing = 'openpyxl' if 'openpyxl' in str(e).lower() else ('xlrd' if 'xlrd' in str(e).lower() else None)
        if missing:
            raise Exception(f"Error parsing Excel file: Missing optional dependency '{missing}'. Install it in the running environment.")
        raise Exception(f"Error parsing Excel file: {str(e)}")
    except Exception as e:
        raise Exception(f"Error parsing Excel file: {str(e)}")


def import_repair_requests_from_dataframe(df):
    """
    Import repair requests from DataFrame.
    Updates existing records if ID already exists.
    """
    imported_count = 0
    updated_count = 0
    errors = []
    
    for index, row in df.iterrows():
        try:
            # Prepare data for model
            data = row.to_dict()
            
            # Remove NaN values and convert to None
            for key, value in data.items():
                if pd.isna(value):
                    data[key] = None
            
            # Get or create the repair request
            repair_request, created = RepairRequest.objects.get_or_create(
                id_field=data['id_field'],
                defaults=data
            )
            
            if not created:
                # Update existing record
                for field, value in data.items():
                    if field != 'id_field':  # Don't update the ID field
                        setattr(repair_request, field, value)
                repair_request.save()
                updated_count += 1
            else:
                imported_count += 1
                
        except Exception as e:
            errors.append(f"Row {index + 1}: {str(e)}")
    
    return {
        'imported': imported_count,
        'updated': updated_count,
        'errors': errors
    }


def create_excel_template():
    """
    Create an Excel template with the correct column headers.
    """
    template_data = {
        'Position': [1],
        'ID': ['REQ-001'],
        'Creator': ['John Doe'],
        'Recipients': ['Technician A'],
        'Location': ['Room 101'],
        'Location path': ['Building A / Floor 1 / Room 101'],
        'Type': ['Plumbing'],
        'Type path': ['Maintenance / Plumbing'],
        'Assets': ['Sink, Faucet'],
        'Ticket': ['TKT-001'],
        'Creation date': [datetime.now()],
        'Priority': ['High'],
        'State': ['Open'],
        'Latest state change user': ['John Doe'],
        'Latest state change time': [datetime.now()],
        'Time accepted': [None],
        'Time in progress': [None],
        'Time done': [None],
        'Time "in evaluation"': [None],
        'Text': ['Sink is leaking'],
        'Link': ['https://example.com'],
        'Submitted result': [None],
        'Comments': [None],
        'Parking reason': [None],
        'Parking information': [None],
    }
    
    df = pd.DataFrame(template_data)
    return df


def get_repair_kpis(start_date=None, end_date=None):
    """
    Calculate KPIs for repair requests.
    """
    queryset = RepairRequest.objects.all()
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    # Calculate average response time
    response_times = queryset.exclude(response_time__isnull=True).values_list('response_time', flat=True)
    if response_times:
        total_response_time = sum(response_times, timezone.timedelta())
        avg_response_time = total_response_time / len(response_times)
    else:
        avg_response_time = None
    
    # Calculate average completion time
    completion_times = queryset.exclude(completion_time__isnull=True).values_list('completion_time', flat=True)
    if completion_times:
        total_completion_time = sum(completion_times, timezone.timedelta())
        avg_completion_time = total_completion_time / len(completion_times)
    else:
        avg_completion_time = None
    
    # Calculate average execution time as (time_done - time_accepted)
    diffs = []
    for td, ta in queryset.filter(time_done__isnull=False, time_accepted__isnull=False).values_list('time_done', 'time_accepted'):
        try:
            diffs.append(td - ta)
        except Exception:
            continue
    if diffs:
        avg_execution_time = sum(diffs, timezone.timedelta()) / len(diffs)
    else:
        avg_execution_time = None
    
    # Count open requests
    open_requests = queryset.filter(state__in=['Open', 'In Progress', 'Accepted', 'In Evaluation']).count()
    
    return {
        'avg_response_time': avg_response_time,
        'avg_completion_time': avg_completion_time,
        'avg_execution_time': avg_execution_time,
        'open_requests': open_requests
    }


def get_repair_trends(start_date=None, end_date=None):
    """
    Get daily trends for repair requests.
    """
    from django.db import models
    
    queryset = RepairRequest.objects.all()
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    # Get created counts by date
    created_data = queryset.extra(
        select={'date': 'DATE(creation_date)'}
    ).values('date').annotate(
        created_count=models.Count('id')
    ).order_by('date')
    
    # Get closed counts by date
    closed_data = queryset.filter(
        state__in=['Closed', 'Done', 'Completed', 'Resolved']
    ).extra(
        select={'date': 'DATE(time_done)'}
    ).values('date').annotate(
        closed_count=models.Count('id')
    ).order_by('date')
    
    # Merge the data
    trend_data = {}
    for item in created_data:
        trend_data[item['date']] = {
            'date': item['date'],
            'created_count': item['created_count'],
            'closed_count': 0
        }
    
    for item in closed_data:
        if item['date'] in trend_data:
            trend_data[item['date']]['closed_count'] = item['closed_count']
        else:
            trend_data[item['date']] = {
                'date': item['date'],
                'created_count': 0,
                'closed_count': item['closed_count']
            }
    
    # Return chronologically sorted list by date
    sorted_items = sorted(trend_data.values(), key=lambda x: x['date'])
    return sorted_items


def get_repair_types(start_date=None, end_date=None):
    """
    Get distribution of repair requests by type.
    """
    from django.db import models
    
    queryset = RepairRequest.objects.all()
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    type_data = queryset.values('type').annotate(
        count=models.Count('id')
    ).order_by('-count')
    
    total_count = sum(item['count'] for item in type_data)
    
    result = []
    for item in type_data:
        percentage = (item['count'] / total_count * 100) if total_count > 0 else 0
        result.append({
            'type': item['type'],
            'count': item['count'],
            'percentage': round(percentage, 2)
        })
    
    return result


def get_repair_heatmap(start_date=None, end_date=None):
    """
    Get heatmap data for repair requests by location.
    """
    from django.db import models
    
    queryset = RepairRequest.objects.all()
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    heatmap_data = queryset.values('location').annotate(
        count=models.Count('id')
    ).order_by('-count')
    
    result = []
    for item in heatmap_data:
        # Extract floor information if available
        floor = None
        if 'floor' in item['location'].lower():
            floor = item['location'].split()[0] if item['location'].split() else None
        
        result.append({
            'location': item['location'],
            'count': item['count'],
            'floor': floor
        })
    
    return result


def get_top_rooms(start_date=None, end_date=None, limit=5):
    """
    Get top rooms by repair request count.
    """
    from django.db import models
    
    queryset = RepairRequest.objects.all()
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    top_rooms = queryset.values('location').annotate(
        count=models.Count('id'),
        avg_completion_time=models.Avg('completion_time')
    ).order_by('-count')[:limit]
    
    return list(top_rooms)


def get_technician_performance(start_date=None, end_date=None):
    """
    Get technician performance data.
    """
    from django.db import models
    
    queryset = RepairRequest.objects.exclude(recipients__isnull=True).exclude(recipients='')
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    tech_data = queryset.values('recipients').annotate(
        count=models.Count('id'),
        avg_response_time=models.Avg('response_time'),
        avg_completion_time=models.Avg('completion_time')
    ).order_by('-count')
    
    return list(tech_data)


def get_sla_compliance(start_date=None, end_date=None):
    """
    Get SLA compliance rates.
    """
    from django.db import models
    
    queryset = RepairRequest.objects.filter(
        state__in=['Closed', 'Done', 'Completed', 'Resolved']
    ).exclude(completion_time__isnull=True)
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    total_count = queryset.count()
    
    if total_count == 0:
        return []
    
    # 4-hour SLA
    compliant_4h = queryset.filter(completion_time__lte=timezone.timedelta(hours=4)).count()
    
    # 24-hour SLA
    compliant_24h = queryset.filter(completion_time__lte=timezone.timedelta(hours=24)).count()
    
    # 48-hour SLA
    compliant_48h = queryset.filter(completion_time__lte=timezone.timedelta(hours=48)).count()
    
    return [
        {
            'sla_period': '4 hours',
            'compliant_count': compliant_4h,
            'total_count': total_count,
            'compliance_rate': round((compliant_4h / total_count) * 100, 2)
        },
        {
            'sla_period': '24 hours',
            'compliant_count': compliant_24h,
            'total_count': total_count,
            'compliance_rate': round((compliant_24h / total_count) * 100, 2)
        },
        {
            'sla_period': '48 hours',
            'compliant_count': compliant_48h,
            'total_count': total_count,
            'compliance_rate': round((compliant_48h / total_count) * 100, 2)
        }
    ]


# Advanced Reporting Functions
def get_sla_compliance_advanced(start_date=None, end_date=None):
    """Get advanced SLA compliance data with more detailed breakdown."""
    from django.db import models
    
    queryset = RepairRequest.objects.filter(
        state__in=['Closed', 'Done', 'Completed', 'Resolved']
    ).exclude(completion_time__isnull=True)
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    total_count = queryset.count()
    
    if total_count == 0:
        return {
            'total_requests': 0,
            'sla_breakdown': [],
            'priority_breakdown': [],
            'trend_data': []
        }
    
    # SLA breakdown by time periods
    sla_breakdown = [
        {
            'period': '1 hour',
            'compliant': queryset.filter(completion_time__lte=timezone.timedelta(hours=1)).count(),
            'total': total_count
        },
        {
            'period': '4 hours',
            'compliant': queryset.filter(completion_time__lte=timezone.timedelta(hours=4)).count(),
            'total': total_count
        },
        {
            'period': '8 hours',
            'compliant': queryset.filter(completion_time__lte=timezone.timedelta(hours=8)).count(),
            'total': total_count
        },
        {
            'period': '24 hours',
            'compliant': queryset.filter(completion_time__lte=timezone.timedelta(hours=24)).count(),
            'total': total_count
        },
        {
            'period': '48 hours',
            'compliant': queryset.filter(completion_time__lte=timezone.timedelta(hours=48)).count(),
            'total': total_count
        }
    ]
    
    # Calculate compliance rates
    for item in sla_breakdown:
        item['compliance_rate'] = round((item['compliant'] / item['total']) * 100, 2) if item['total'] > 0 else 0
    
    # Priority breakdown
    priority_breakdown = queryset.values('priority').annotate(
        count=models.Count('id'),
        avg_completion=models.Avg('completion_time'),
        sla_4h_compliant=models.Count('id', filter=models.Q(completion_time__lte=timezone.timedelta(hours=4))),
        sla_24h_compliant=models.Count('id', filter=models.Q(completion_time__lte=timezone.timedelta(hours=24)))
    ).order_by('-count')
    
    for item in priority_breakdown:
        item['sla_4h_rate'] = round((item['sla_4h_compliant'] / item['count']) * 100, 2) if item['count'] > 0 else 0
        item['sla_24h_rate'] = round((item['sla_24h_compliant'] / item['count']) * 100, 2) if item['count'] > 0 else 0
    
    return {
        'total_requests': total_count,
        'sla_breakdown': sla_breakdown,
        'priority_breakdown': list(priority_breakdown)
    }


def get_delay_by_priority(start_date=None, end_date=None):
    """Get average response/completion time grouped by priority."""
    from django.db import models
    
    queryset = RepairRequest.objects.all()
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    priority_data = queryset.exclude(priority__isnull=True).exclude(priority='').values('priority').annotate(
        count=models.Count('id'),
        avg_response_time=models.Avg('response_time'),
        avg_completion_time=models.Avg('completion_time'),
        avg_execution_time=models.Avg('execution_time'),
        avg_evaluation_time=models.Avg('evaluation_time')
    ).order_by('-count')
    
    return list(priority_data)


def get_escalations(start_date=None, end_date=None):
    """Get count of requests with changed recipients (escalations)."""
    from django.db import models
    
    queryset = RepairRequest.objects.all()
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    # This is a simplified version - in a real system you'd track recipient changes
    # For now, we'll count requests where recipients contain multiple people (indicating escalation)
    escalated_requests = queryset.exclude(recipients__isnull=True).exclude(recipients='').filter(
        recipients__contains=','
    ).count()
    
    total_requests = queryset.count()
    
    return {
        'escalated_requests': escalated_requests,
        'total_requests': total_requests,
        'escalation_rate': round((escalated_requests / total_requests) * 100, 2) if total_requests > 0 else 0
    }


def get_technician_performance_advanced(start_date=None, end_date=None):
    """Get advanced technician performance data."""
    from django.db import models
    
    queryset = RepairRequest.objects.exclude(recipients__isnull=True).exclude(recipients='')
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    # Split recipients and get individual technician performance
    technician_data = {}
    
    for repair in queryset:
        recipients = repair.recipients.split(',') if repair.recipients else []
        for recipient in recipients:
            recipient = recipient.strip()
            if recipient not in technician_data:
                technician_data[recipient] = {
                    'technician': recipient,
                    'total_requests': 0,
                    'completed_requests': 0,
                    'response_times': [],
                    'completion_times': [],
                    'execution_times': []
                }
            
            technician_data[recipient]['total_requests'] += 1
            
            if repair.is_closed:
                technician_data[recipient]['completed_requests'] += 1
            
            if repair.response_time:
                technician_data[recipient]['response_times'].append(repair.response_time)
            
            if repair.completion_time:
                technician_data[recipient]['completion_times'].append(repair.completion_time)
            
            if repair.execution_time:
                technician_data[recipient]['execution_times'].append(repair.execution_time)
    
    # Calculate averages
    result = []
    for tech_data in technician_data.values():
        response_times = tech_data['response_times']
        completion_times = tech_data['completion_times']
        execution_times = tech_data['execution_times']
        
        avg_response = sum(response_times, timezone.timedelta()) / len(response_times) if response_times else None
        avg_completion = sum(completion_times, timezone.timedelta()) / len(completion_times) if completion_times else None
        avg_execution = sum(execution_times, timezone.timedelta()) / len(execution_times) if execution_times else None
        
        completion_rate = (tech_data['completed_requests'] / tech_data['total_requests']) * 100 if tech_data['total_requests'] > 0 else 0
        
        result.append({
            'technician': tech_data['technician'],
            'total_requests': tech_data['total_requests'],
            'completed_requests': tech_data['completed_requests'],
            'completion_rate': round(completion_rate, 2),
            'avg_response_time': avg_response,
            'avg_completion_time': avg_completion,
            'avg_execution_time': avg_execution
        })
    
    return sorted(result, key=lambda x: x['total_requests'], reverse=True)


def get_reopened_requests(start_date=None, end_date=None):
    """Get requests that were closed then reopened."""
    from django.db import models
    
    queryset = RepairRequest.objects.all()
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    # This is a simplified version - in a real system you'd track state changes
    # For now, we'll look for requests that have been in evaluation after being done
    reopened_requests = queryset.filter(
        time_done__isnull=False,
        time_in_evaluation__isnull=False,
        time_in_evaluation__gt=models.F('time_done')
    ).count()
    
    total_closed = queryset.filter(
        state__in=['Closed', 'Done', 'Completed', 'Resolved']
    ).count()
    
    return {
        'reopened_requests': reopened_requests,
        'total_closed': total_closed,
        'reopening_rate': round((reopened_requests / total_closed) * 100, 2) if total_closed > 0 else 0
    }


def get_workload_distribution():
    """Get current open requests per technician."""
    from django.db import models
    
    queryset = RepairRequest.objects.filter(
        state__in=['Open', 'In Progress', 'Accepted', 'In Evaluation']
    ).exclude(recipients__isnull=True).exclude(recipients='')
    
    technician_workload = {}
    
    for repair in queryset:
        recipients = repair.recipients.split(',') if repair.recipients else []
        for recipient in recipients:
            recipient = recipient.strip()
            if recipient not in technician_workload:
                technician_workload[recipient] = 0
            technician_workload[recipient] += 1
    
    result = []
    for technician, count in technician_workload.items():
        result.append({
            'technician': technician,
            'open_requests': count
        })
    
    return sorted(result, key=lambda x: x['open_requests'], reverse=True)


def get_top_assets(start_date=None, end_date=None, limit=5):
    """Get top assets by repair request count."""
    from django.db import models
    
    queryset = RepairRequest.objects.exclude(assets__isnull=True).exclude(assets='')
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    asset_data = {}
    
    for repair in queryset:
        assets = repair.assets.split(',') if repair.assets else []
        for asset in assets:
            asset = asset.strip()
            if asset not in asset_data:
                asset_data[asset] = {
                    'asset': asset,
                    'count': 0,
                    'avg_completion_time': [],
                    'locations': set()
                }
            
            asset_data[asset]['count'] += 1
            asset_data[asset]['locations'].add(repair.location)
            
            if repair.completion_time:
                asset_data[asset]['avg_completion_time'].append(repair.completion_time)
    
    # Calculate averages and convert sets to lists
    result = []
    for asset_info in asset_data.values():
        completion_times = asset_info['avg_completion_time']
        avg_completion = sum(completion_times, timezone.timedelta()) / len(completion_times) if completion_times else None
        
        result.append({
            'asset': asset_info['asset'],
            'count': asset_info['count'],
            'avg_completion_time': avg_completion,
            'locations': list(asset_info['locations']),
            'location_count': len(asset_info['locations'])
        })
    
    return sorted(result, key=lambda x: x['count'], reverse=True)[:limit]


def get_repeat_requests(start_date=None, end_date=None):
    """Get repeat requests for same location/type."""
    from django.db import models
    
    queryset = RepairRequest.objects.all()
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    # Group by location and type
    location_type_groups = queryset.values('location', 'type').annotate(
        count=models.Count('id')
    ).filter(count__gt=1).order_by('-count')
    
    result = []
    for group in location_type_groups:
        # Get all requests for this location/type combination
        requests = queryset.filter(
            location=group['location'],
            type=group['type']
        ).order_by('creation_date')
        
        result.append({
            'location': group['location'],
            'type': group['type'],
            'count': group['count'],
            'first_request': requests.first().creation_date if requests.exists() else None,
            'last_request': requests.last().creation_date if requests.exists() else None,
            'avg_time_between': None  # Could calculate this if needed
        })
    
    return result


def get_bottlenecks(start_date=None, end_date=None):
    """Get process bottlenecks analysis."""
    from django.db import models
    
    queryset = RepairRequest.objects.all()
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    # Analyze time spent in each state
    state_analysis = queryset.values('state').annotate(
        count=models.Count('id'),
        avg_response_time=models.Avg('response_time'),
        avg_execution_time=models.Avg('execution_time'),
        avg_evaluation_time=models.Avg('evaluation_time')
    ).order_by('-count')
    
    # Find bottlenecks (states with longest average times)
    bottlenecks = []
    for state in state_analysis:
        if state['avg_response_time'] or state['avg_execution_time'] or state['avg_evaluation_time']:
            bottlenecks.append({
                'state': state['state'],
                'count': state['count'],
                'avg_response_time': state['avg_response_time'],
                'avg_execution_time': state['avg_execution_time'],
                'avg_evaluation_time': state['avg_evaluation_time']
            })
    
    return sorted(bottlenecks, key=lambda x: x['count'], reverse=True)


def get_avg_evaluation_time(start_date=None, end_date=None):
    """Get average evaluation time."""
    from django.db import models
    
    queryset = RepairRequest.objects.exclude(evaluation_time__isnull=True)
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    evaluation_times = queryset.values_list('evaluation_time', flat=True)
    
    if evaluation_times:
        total_evaluation_time = sum(evaluation_times, timezone.timedelta())
        avg_evaluation_time = total_evaluation_time / len(evaluation_times)
        
        return {
            'avg_evaluation_time': avg_evaluation_time,
            'total_requests_in_evaluation': len(evaluation_times),
            'min_evaluation_time': min(evaluation_times),
            'max_evaluation_time': max(evaluation_times)
        }
    
    return {
        'avg_evaluation_time': None,
        'total_requests_in_evaluation': 0,
        'min_evaluation_time': None,
        'max_evaluation_time': None
    }


def get_parking_reasons(start_date=None, end_date=None):
    """Get parking reasons analysis."""
    from django.db import models
    
    queryset = RepairRequest.objects.exclude(parking_reason__isnull=True).exclude(parking_reason='')
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    parking_reasons = queryset.values('parking_reason').annotate(
        count=models.Count('id')
    ).order_by('-count')
    
    return list(parking_reasons)


def get_guest_facing_requests(start_date=None, end_date=None):
    """Get guest-facing requests analysis."""
    from django.db import models
    
    queryset = RepairRequest.objects.all()
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    # This is a simplified version - in a real system you'd have a field to identify guest-facing requests
    # For now, we'll assume requests in guest rooms are guest-facing
    guest_facing = queryset.filter(
        location__icontains='room'
    ).exclude(location__icontains='staff')
    
    total_requests = queryset.count()
    guest_facing_count = guest_facing.count()
    
    return {
        'guest_facing_requests': guest_facing_count,
        'total_requests': total_requests,
        'guest_facing_percentage': round((guest_facing_count / total_requests) * 100, 2) if total_requests > 0 else 0,
        'avg_completion_time_guest': guest_facing.aggregate(
            avg=models.Avg('completion_time')
        )['avg']
    }


def get_internal_requests(start_date=None, end_date=None):
    """Get internal requests analysis."""
    from django.db import models
    
    queryset = RepairRequest.objects.all()
    
    if start_date:
        queryset = queryset.filter(creation_date__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(creation_date__date__lte=end_date)
    
    # This is a simplified version - in a real system you'd have a field to identify internal requests
    # For now, we'll assume requests not in guest rooms are internal
    internal = queryset.exclude(location__icontains='room')
    
    total_requests = queryset.count()
    internal_count = internal.count()
    
    return {
        'internal_requests': internal_count,
        'total_requests': total_requests,
        'internal_percentage': round((internal_count / total_requests) * 100, 2) if total_requests > 0 else 0,
        'avg_completion_time_internal': internal.aggregate(
            avg=models.Avg('completion_time')
        )['avg']
    }


# Export Functions
def export_to_excel(report_type, start_date=None, end_date=None):
    """Export report data to Excel."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    import io
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{report_type.title()} Report"
    
    # Add header
    ws['A1'] = f"{report_type.title()} Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Date Range: {start_date or 'All'} to {end_date or 'All'}"
    ws['A2'].font = Font(size=12)
    
    # Get data based on report type
    if report_type == 'daily':
        data = get_daily_flash_data(start_date, end_date)
    elif report_type == 'weekly':
        data = get_weekly_trend_data(start_date, end_date)
    elif report_type == 'monthly':
        data = get_monthly_root_cause_data(start_date, end_date)
    else:
        data = {}
    
    # Add data to worksheet
    row = 4
    for section, section_data in data.items():
        ws[f'A{row}'] = section
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        if isinstance(section_data, list) and section_data:
            # Add headers
            headers = list(section_data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            row += 1
            
            # Add data rows
            for item in section_data:
                for col, header in enumerate(headers, 1):
                    value = item.get(header, '')
                    if hasattr(value, 'total_seconds'):  # Handle timedelta objects
                        value = str(value)
                    ws.cell(row=row, column=col, value=value)
                row += 1
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.read()


def export_to_pdf(report_type, start_date=None, end_date=None):
    """Export report data to PDF."""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    import io
    
    # Create BytesIO buffer
    buffer = io.BytesIO()
    
    # Create PDF document
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Get data based on report type
    if report_type == 'daily':
        data = get_daily_flash_data(start_date, end_date)
    elif report_type == 'weekly':
        data = get_weekly_trend_data(start_date, end_date)
    elif report_type == 'monthly':
        data = get_monthly_root_cause_data(start_date, end_date)
    else:
        data = {}
    
    # Build PDF content
    story = []
    
    # Add title
    story.append(Paragraph(f"{report_type.title()} Report", title_style))
    story.append(Paragraph(f"Date Range: {start_date or 'All'} to {end_date or 'All'}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Add data sections
    for section, section_data in data.items():
        story.append(Paragraph(section, styles['Heading2']))
        story.append(Spacer(1, 12))
        
        if isinstance(section_data, list) and section_data:
            # Create table
            headers = list(section_data[0].keys())
            table_data = [headers]
            
            for item in section_data:
                row = []
                for header in headers:
                    value = item.get(header, '')
                    if hasattr(value, 'total_seconds'):  # Handle timedelta objects
                        value = str(value)
                    row.append(str(value))
                table_data.append(row)
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
    
    # Build PDF
    doc.build(story)
    
    # Get PDF content
    buffer.seek(0)
    return buffer.read()


def get_daily_flash_data(start_date=None, end_date=None):
    """Get data for daily flash report."""
    kpis = get_repair_kpis(start_date, end_date)
    top_rooms = get_top_rooms(start_date, end_date, 5)
    technicians = get_technician_performance_advanced(start_date, end_date)[:5]
    sla_data = get_sla_compliance_advanced(start_date, end_date)
    
    return {
        'KPIs': [kpis],
        'Top Rooms': top_rooms,
        'Top Technicians': technicians,
        'SLA Compliance': sla_data['sla_breakdown']
    }


def get_weekly_trend_data(start_date=None, end_date=None):
    """Get data for weekly trend report."""
    trends = get_repair_trends(start_date, end_date)
    sla_trends = get_sla_compliance_advanced(start_date, end_date)
    types = get_repair_types(start_date, end_date)
    
    return {
        'Daily Trends': trends,
        'SLA Trends': sla_trends['sla_breakdown'],
        'Type Distribution': types
    }


def get_monthly_root_cause_data(start_date=None, end_date=None):
    """Get data for monthly root cause report."""
    types = get_repair_types(start_date, end_date)[:5]
    locations = get_repair_heatmap(start_date, end_date)[:5]
    sla_breakdown = get_sla_compliance_advanced(start_date, end_date)
    bottlenecks = get_bottlenecks(start_date, end_date)
    
    return {
        'Top 5 Request Types': types,
        'Top 5 Locations': locations,
        'SLA Breakdown': sla_breakdown['sla_breakdown'],
        'Process Bottlenecks': bottlenecks
    }